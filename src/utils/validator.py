"""Validador de integridade do pipeline ETL financeiro."""

import sys
from collections import Counter
from pathlib import Path

import openpyxl

from src.utils.logger import configurar_logger

logger = configurar_logger("validator")

RAIZ = Path(__file__).parent.parent.parent
DIR_OUTPUT = RAIZ / "data" / "output"

CLASSIFICACOES_VALIDAS: set[str] = {
    "Obrigatório",
    "Questionável",
    "Supérfluo",
    "N/A",
}


def _encontrar_xlsx_mais_recente() -> Path | None:
    """Encontra o XLSX de controle de bordo mais recente em data/output/."""
    arquivos = sorted(DIR_OUTPUT.glob("controle_bordo_*.xlsx"), reverse=True)
    if not arquivos:
        return None
    return arquivos[0]


def _carregar_transacoes(caminho: Path) -> list[dict]:
    """Carrega transações da aba 'extrato' do XLSX."""
    wb = openpyxl.load_workbook(caminho)
    if "extrato" not in wb.sheetnames:
        logger.error("Aba 'extrato' não encontrada em %s", caminho)
        return []

    ws = wb["extrato"]
    headers = [cell.value for cell in ws[1]]
    transacoes: list[dict] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        registro: dict = {}
        for idx, valor in enumerate(row):
            if idx < len(headers):
                registro[headers[idx]] = valor
        if registro.get("data") is not None:
            transacoes.append(registro)

    return transacoes


def validar_total_por_banco(transacoes: list[dict]) -> tuple[bool, str]:
    """Verifica total de transações por banco (consistência)."""
    contagem: Counter[str] = Counter()
    for t in transacoes:
        banco = t.get("banco_origem", "Desconhecido")
        contagem[banco] += 1

    detalhes = ", ".join(f"{banco}({qtd})" for banco, qtd in contagem.most_common())
    return True, f"{len(transacoes)} transações processadas -- {detalhes}"


def validar_sem_categoria(transacoes: list[dict]) -> tuple[bool, str]:
    """Verifica se existem transações sem categoria."""
    sem_categoria = [
        t for t in transacoes
        if t.get("categoria") is None or t.get("categoria") == ""
    ]
    qtd = len(sem_categoria)
    ok = qtd == 0
    return ok, f"{qtd} transações sem categoria"


def validar_classificacoes(transacoes: list[dict]) -> tuple[bool, str]:
    """Verifica se todas as classificações são válidas."""
    contagem: Counter[str] = Counter()
    invalidas: list[str] = []

    for t in transacoes:
        clf = t.get("classificacao", "")
        if clf in CLASSIFICACOES_VALIDAS:
            contagem[clf] += 1
        else:
            invalidas.append(str(clf))

    detalhes = ", ".join(
        f"{clf}({contagem[clf]})" for clf in CLASSIFICACOES_VALIDAS if contagem[clf]
    )

    if invalidas:
        contagem_inv: Counter[str] = Counter(invalidas)
        detalhes_inv = ", ".join(f"{v}({c})" for v, c in contagem_inv.most_common(5))
        return False, f"Classificações inválidas encontradas: {detalhes_inv}"

    return True, f"Classificações válidas: {detalhes}"


def validar_duplicatas(transacoes: list[dict]) -> tuple[bool, str]:
    """Verifica duplicatas residuais (mesma data + valor + local)."""
    chaves: Counter[str] = Counter()
    for t in transacoes:
        data = t.get("data", "")
        data_str = data.isoformat() if hasattr(data, "isoformat") else str(data)
        valor = t.get("valor", 0)
        local = t.get("local", "")
        chave = f"{data_str}|{valor}|{local}"
        chaves[chave] += 1

    duplicatas = {k: v for k, v in chaves.items() if v > 1}
    total_dups = sum(v - 1 for v in duplicatas.values())

    if total_dups > 0:
        return False, f"{total_dups} duplicatas residuais encontradas"

    return True, "Sem duplicatas residuais"


def validar_meses_receita_zero(transacoes: list[dict]) -> tuple[bool, str]:
    """Verifica meses com receita zero (possível falha de extração)."""
    meses: set[str] = set()
    receita_por_mes: dict[str, float] = {}

    for t in transacoes:
        mes = t.get("mes_ref", "")
        if not mes:
            continue
        meses.add(mes)
        if t.get("tipo") == "Receita":
            receita_por_mes[mes] = receita_por_mes.get(mes, 0) + t.get("valor", 0)

    meses_sem_receita = sorted(m for m in meses if receita_por_mes.get(m, 0) == 0)

    if meses_sem_receita:
        return False, (
            f"{len(meses_sem_receita)} meses com receita zero "
            f"({', '.join(meses_sem_receita)}) -- dados históricos sem renda"
        )

    return True, "Todos os meses possuem receita"


def validar_despesa_maior_receita(transacoes: list[dict]) -> tuple[bool, str]:
    """Alerta se despesa > receita em algum mês."""
    dados_mes: dict[str, dict[str, float]] = {}

    for t in transacoes:
        if t.get("tipo") == "Transferência Interna":
            continue

        mes = t.get("mes_ref", "")
        if not mes:
            continue

        if mes not in dados_mes:
            dados_mes[mes] = {"receita": 0.0, "despesa": 0.0}

        valor = t.get("valor", 0)
        if t.get("tipo") == "Receita":
            dados_mes[mes]["receita"] += valor
        elif t.get("tipo") in ("Despesa", "Imposto"):
            dados_mes[mes]["despesa"] += valor

    meses_negativos = sorted(
        m for m, d in dados_mes.items()
        if d["receita"] > 0 and d["despesa"] > d["receita"]
    )

    if meses_negativos:
        return False, (
            f"{len(meses_negativos)} meses com despesa > receita "
            f"({', '.join(meses_negativos)})"
        )

    return True, "Nenhum mês com despesa superior à receita"


def executar_validacao(caminho: Path | None = None) -> bool:
    """Executa todas as validações e imprime resultado formatado."""
    if caminho is None:
        caminho = _encontrar_xlsx_mais_recente()

    if caminho is None or not caminho.exists():
        logger.error("Nenhum XLSX de controle de bordo encontrado em %s", DIR_OUTPUT)
        return False

    logger.info("Validando: %s", caminho)

    transacoes = _carregar_transacoes(caminho)
    if not transacoes:
        logger.error("Nenhuma transação encontrada no XLSX")
        return False

    validacoes: list[tuple[str, callable]] = [
        ("Total por banco", validar_total_por_banco),
        ("Transações sem categoria", validar_sem_categoria),
        ("Classificações válidas", validar_classificacoes),
        ("Duplicatas residuais", validar_duplicatas),
        ("Meses com receita zero", validar_meses_receita_zero),
        ("Despesa > Receita", validar_despesa_maior_receita),
    ]

    todas_ok = True

    logger.info("=== Validação de Integridade ===")
    for _nome, func in validacoes:
        ok, mensagem = func(transacoes)
        prefixo = "[OK]" if ok else "[!]"
        logger.info("%s %s", prefixo, mensagem)
        if not ok:
            todas_ok = False

    return todas_ok


def main() -> None:
    """Entrypoint CLI para validação."""
    resultado = executar_validacao()
    sys.exit(0 if resultado else 1)


if __name__ == "__main__":
    main()


# "A medida do homem e o que ele faz com o poder que possui." -- Platao
