"""Detecção automática de banco, tipo, pessoa e período de arquivos financeiros pelo conteúdo."""

import csv
import hashlib
import io
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import msoffcrypto
import openpyxl
import pdfplumber
import xlrd

from src.utils.logger import configurar_logger
from src.utils.senhas import carregar_senhas_pdf

logger = configurar_logger("file_detector")

RAIZ_PROJETO = Path(__file__).parent.parent.parent

MESES_POR_EXTENSO: dict[str, str] = {
    "janeiro": "01",
    "fevereiro": "02",
    "marco": "03",
    "março": "03",
    "abril": "04",
    "maio": "05",
    "junho": "06",
    "julho": "07",
    "agosto": "08",
    "setembro": "09",
    "outubro": "10",
    "novembro": "11",
    "dezembro": "12",
    "jan": "01",
    "fev": "02",
    "mar": "03",
    "abr": "04",
    "mai": "05",
    "jun": "06",
    "jul": "07",
    "ago": "08",
    "set": "09",
    "out": "10",
    "nov": "11",
    "dez": "12",
}

MESES_ABREV_UPPER: dict[str, str] = {
    "JAN": "01",
    "FEV": "02",
    "MAR": "03",
    "ABR": "04",
    "MAI": "05",
    "JUN": "06",
    "JUL": "07",
    "AGO": "08",
    "SET": "09",
    "OUT": "10",
    "NOV": "11",
    "DEZ": "12",
}


@dataclass
class DeteccaoArquivo:
    """Resultado da detecção automática de um arquivo financeiro."""

    banco: str
    tipo: str
    pessoa: str
    subtipo: str
    periodo: Optional[str]
    formato: str
    confianca: float


def _carregar_senhas() -> list[str]:
    """Carrega senhas via módulo centralizado."""
    return carregar_senhas_pdf()


def calcular_hash(caminho: Path) -> str:
    """Calcula SHA-256 de um arquivo para detecção de duplicatas."""
    sha256 = hashlib.sha256()
    with open(caminho, "rb") as f:
        for bloco in iter(lambda: f.read(8192), b""):
            sha256.update(bloco)
    return sha256.hexdigest()


def _extrair_periodo_csv_nubank_cartao(caminho: Path) -> Optional[str]:
    """Extrai período de CSV de cartão Nubank pela coluna date."""
    try:
        with open(caminho, encoding="utf-8") as f:
            leitor = csv.DictReader(f)
            datas: list[str] = []
            for linha in leitor:
                data = linha.get("date", "").strip()
                if data:
                    datas.append(data)
            if datas:
                primeira = datas[0]
                match = re.match(r"(\d{4})-(\d{2})", primeira)
                if match:
                    return f"{match.group(1)}-{match.group(2)}"
    except Exception as erro:
        logger.debug("Erro ao extrair período de CSV cartão: %s", erro)
    return None


def _extrair_periodo_csv_nubank_cc(caminho: Path) -> Optional[str]:
    """Extrai período de CSV de conta corrente Nubank pela coluna Data."""
    try:
        with open(caminho, encoding="utf-8") as f:
            leitor = csv.DictReader(f)
            datas: list[str] = []
            for linha in leitor:
                data = linha.get("Data", "").strip()
                if data:
                    datas.append(data)
            if datas:
                primeira = datas[0]
                match = re.match(r"(\d{2})/(\d{2})/(\d{4})", primeira)
                if match:
                    return f"{match.group(3)}-{match.group(2)}"
    except Exception as erro:
        logger.debug("Erro ao extrair período de CSV CC: %s", erro)
    return None


def _extrair_periodo_do_nome_nu(nome: str) -> Optional[str]:
    """Extrai período do nome de arquivo padrão NU_977370681_01MES2025_31MES2025.csv."""
    match = re.search(r"(\d{2})([A-Z]{3})(\d{4})", nome.upper())
    if match:
        mes_abrev = match.group(2)
        ano = match.group(3)
        mes_num = MESES_ABREV_UPPER.get(mes_abrev)
        if mes_num:
            return f"{ano}-{mes_num}"
    return None


def _extrair_periodo_xls_fatura(nome: str) -> Optional[str]:
    """Extrai período do nome do arquivo XLS tipo Fatura-CPF-{mes}.xls."""
    nome_lower = nome.lower()
    for mes_nome, mes_num in MESES_POR_EXTENSO.items():
        if mes_nome in nome_lower:
            match_ano = re.search(r"(\d{4})", nome)
            ano = match_ano.group(1) if match_ano else "2026"
            return f"{ano}-{mes_num}"
    return None


def _detectar_csv(caminho: Path) -> Optional[DeteccaoArquivo]:
    """Detecta banco, tipo e pessoa para arquivos CSV."""
    nome = caminho.name
    caminho_str = str(caminho).lower()

    try:
        with open(caminho, encoding="utf-8") as f:
            primeira_linha = f.readline().strip()
            conteudo = f.read()
    except Exception as erro:
        logger.warning("Erro ao ler CSV %s: %s", caminho, erro)
        return None

    conteudo_completo = primeira_linha + "\n" + conteudo

    if primeira_linha == "date,title,amount":
        periodo = _extrair_periodo_csv_nubank_cartao(caminho)

        if "pj" in caminho_str or (re.search(r"Nubank_202[56]", nome) and "pj" in caminho_str):
            return DeteccaoArquivo(
                banco="nubank",
                tipo="cartao",
                pessoa="vitoria",
                subtipo="pj",
                periodo=periodo,
                formato="csv",
                confianca=0.9,
            )

        if "andre" in caminho_str:
            return DeteccaoArquivo(
                banco="nubank",
                tipo="cartao",
                pessoa="andre",
                subtipo="",
                periodo=periodo,
                formato="csv",
                confianca=0.9,
            )

        if "vitoria" in caminho_str:
            return DeteccaoArquivo(
                banco="nubank",
                tipo="cartao",
                pessoa="vitoria",
                subtipo="pj",
                periodo=periodo,
                formato="csv",
                confianca=0.85,
            )

        return DeteccaoArquivo(
            banco="nubank",
            tipo="cartao",
            pessoa="andre",
            subtipo="",
            periodo=periodo,
            formato="csv",
            confianca=0.7,
        )

    if "Data,Valor,Identificador" in primeira_linha:
        periodo = _extrair_periodo_csv_nubank_cc(caminho)

        if "NU_977370681" in nome or "977370681" in nome:
            periodo_nome = _extrair_periodo_do_nome_nu(nome)
            if periodo_nome:
                periodo = periodo_nome
            return DeteccaoArquivo(
                banco="nubank",
                tipo="cc",
                pessoa="vitoria",
                subtipo="pf",
                periodo=periodo,
                formato="csv",
                confianca=0.95,
            )

        if "cc_pj" in nome.lower() or "pj_cc" in caminho_str:
            return DeteccaoArquivo(
                banco="nubank",
                tipo="cc",
                pessoa="vitoria",
                subtipo="pj",
                periodo=periodo,
                formato="csv",
                confianca=0.9,
            )

        if "52.488.753" in conteudo_completo:
            return DeteccaoArquivo(
                banco="nubank",
                tipo="cc",
                pessoa="vitoria",
                subtipo="pj",
                periodo=periodo,
                formato="csv",
                confianca=0.95,
            )

        if "070.475" in conteudo_completo:
            return DeteccaoArquivo(
                banco="nubank",
                tipo="cc",
                pessoa="vitoria",
                subtipo="pf",
                periodo=periodo,
                formato="csv",
                confianca=0.9,
            )

        if "pj" in caminho_str:
            return DeteccaoArquivo(
                banco="nubank",
                tipo="cc",
                pessoa="vitoria",
                subtipo="pj",
                periodo=periodo,
                formato="csv",
                confianca=0.8,
            )

        if "vitoria" in caminho_str:
            return DeteccaoArquivo(
                banco="nubank",
                tipo="cc",
                pessoa="vitoria",
                subtipo="pf",
                periodo=periodo,
                formato="csv",
                confianca=0.7,
            )

        return DeteccaoArquivo(
            banco="nubank",
            tipo="cc",
            pessoa="andre",
            subtipo="",
            periodo=periodo,
            formato="csv",
            confianca=0.5,
        )

    logger.debug("CSV com headers não reconhecidos: %s", primeira_linha)
    return None


def _abrir_xlsx_decriptado(caminho: Path, senhas: list[str]) -> Optional[openpyxl.Workbook]:
    """Tenta abrir XLSX, decriptando se necessário."""
    try:
        with open(caminho, "rb") as f:
            ms = msoffcrypto.OfficeFile(f)
            if ms.is_encrypted():
                for senha in senhas:
                    try:
                        f.seek(0)
                        ms = msoffcrypto.OfficeFile(f)
                        ms.load_key(password=senha)
                        decriptado = io.BytesIO()
                        ms.decrypt(decriptado)
                        decriptado.seek(0)
                        return openpyxl.load_workbook(decriptado)
                    except Exception:
                        continue
                logger.warning("Nenhuma senha funcionou para XLSX: %s", caminho)
                return None
            else:
                return openpyxl.load_workbook(caminho)
    except Exception:
        try:
            return openpyxl.load_workbook(caminho)
        except Exception as erro:
            logger.warning("Falha ao abrir XLSX %s: %s", caminho, erro)
            return None


def _detectar_xlsx(caminho: Path) -> Optional[DeteccaoArquivo]:
    """Detecta banco, tipo e pessoa para arquivos XLSX."""
    senhas = _carregar_senhas()
    wb = _abrir_xlsx_decriptado(caminho, senhas)
    if wb is None:
        return None

    try:
        ws = wb.active
        conteudo_celulas: list[str] = []
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
            for celula in row:
                if celula is not None:
                    conteudo_celulas.append(str(celula))

        texto_concatenado = " ".join(conteudo_celulas).upper()

        if "C6 BANK" in texto_concatenado or "EXTRATO DE CONTA CORRENTE C6" in texto_concatenado:
            periodo: Optional[str] = None
            for texto in conteudo_celulas:
                match = re.search(
                    r"Extrato de (\d{2}/\d{2}/\d{4}) a (\d{2}/\d{2}/\d{4})",
                    texto,
                    re.IGNORECASE,
                )
                if match:
                    data_fim = match.group(2)
                    partes = data_fim.split("/")
                    periodo = f"{partes[2]}-{partes[1]}"
                    break

            return DeteccaoArquivo(
                banco="c6",
                tipo="cc",
                pessoa="andre",
                subtipo="",
                periodo=periodo,
                formato="xlsx",
                confianca=0.95,
            )
    except Exception as erro:
        logger.warning("Erro ao analisar XLSX %s: %s", caminho, erro)
    finally:
        wb.close()

    return None


def _abrir_xls_decriptado(caminho: Path, senhas: list[str]) -> Optional[xlrd.book.Book]:
    """Tenta abrir XLS (formato legado), decriptando se necessário."""
    for senha in senhas:
        try:
            with open(caminho, "rb") as f:
                ms = msoffcrypto.OfficeFile(f)
                ms.load_key(password=senha)
                decriptado = io.BytesIO()
                ms.decrypt(decriptado)
                decriptado.seek(0)
                return xlrd.open_workbook(file_contents=decriptado.read())
        except Exception:
            continue

    try:
        return xlrd.open_workbook(str(caminho))
    except Exception as erro:
        logger.warning("Falha ao abrir XLS %s: %s", caminho, erro)
        return None


def _detectar_xls(caminho: Path) -> Optional[DeteccaoArquivo]:
    """Detecta banco, tipo e pessoa para arquivos XLS."""
    senhas = _carregar_senhas()
    wb = _abrir_xls_decriptado(caminho, senhas)
    if wb is None:
        return None

    try:
        ws = wb.sheet_by_index(0)
        conteudo_celulas: list[str] = []
        for i in range(min(15, ws.nrows)):
            valores = ws.row_values(i)
            for v in valores:
                if v:
                    conteudo_celulas.append(str(v))

        texto_concatenado = " ".join(conteudo_celulas)

        if "Cartão C6" in texto_concatenado or "C6" in texto_concatenado.upper():
            periodo = _extrair_periodo_xls_fatura(caminho.name)
            return DeteccaoArquivo(
                banco="c6",
                tipo="cartao",
                pessoa="andre",
                subtipo="",
                periodo=periodo,
                formato="xls",
                confianca=0.9,
            )
    except Exception as erro:
        logger.warning("Erro ao analisar XLS %s: %s", caminho, erro)

    return None


def _abrir_pdf(caminho: Path, senhas: list[str]) -> Optional[pdfplumber.PDF]:
    """Tenta abrir PDF, testando senhas se necessário."""
    try:
        pdf = pdfplumber.open(caminho)
        if pdf.pages:
            _ = pdf.pages[0].extract_text()
        return pdf
    except Exception as err:
        logger.debug("PDF %s exige senha ou falhou ao abrir sem senha: %s", caminho.name, err)

    for senha in senhas:
        try:
            pdf = pdfplumber.open(caminho, password=senha)
            if pdf.pages:
                _ = pdf.pages[0].extract_text()
            return pdf
        except Exception:
            continue

    logger.warning("Nenhuma senha funcionou para PDF: %s", caminho)
    return None


def _extrair_periodo_itau(texto: str) -> Optional[str]:
    """Extrai período do extrato Itaú a partir do texto do PDF."""
    match = re.search(
        r"período de visualização:\s*(\d{2}/\d{2}/\d{4})\s*até\s*(\d{2}/\d{2}/\d{4})",
        texto,
        re.IGNORECASE,
    )
    if match:
        data_inicio = match.group(1)
        partes = data_inicio.split("/")
        return f"{partes[2]}-{partes[1]}"
    return None


def _extrair_periodo_santander(texto: str) -> Optional[str]:
    """Extrai período da fatura Santander a partir do texto do PDF."""
    match = re.search(r"Vencimento\s*\n?\s*(\d{2}/\d{2}/\d{4})", texto)
    if match:
        data_venc = match.group(1)
        partes = data_venc.split("/")
        return f"{partes[2]}-{partes[1]}"

    match_curto = re.search(r"(\d{2})/(\w{3})/(\d{2})", texto)
    if match_curto:
        mes_abrev = match_curto.group(2).lower()
        ano_curto = match_curto.group(3)
        mes_num = MESES_POR_EXTENSO.get(mes_abrev)
        if mes_num:
            ano = f"20{ano_curto}"
            return f"{ano}-{mes_num}"

    return None


def _detectar_pdf(caminho: Path) -> Optional[DeteccaoArquivo]:
    """Detecta banco, tipo e pessoa para arquivos PDF."""
    senhas = _carregar_senhas()
    pdf = _abrir_pdf(caminho, senhas)
    if pdf is None:
        return None

    try:
        texto_completo = ""
        for pagina in pdf.pages[:3]:
            texto_pagina = pagina.extract_text()
            if texto_pagina:
                texto_completo += texto_pagina + "\n"

        texto_upper = texto_completo.upper()

        if "ITAÚ UNIBANCO" in texto_upper or "agência: 6450" in texto_completo:
            periodo = _extrair_periodo_itau(texto_completo)
            return DeteccaoArquivo(
                banco="itau",
                tipo="cc",
                pessoa="andre",
                subtipo="",
                periodo=periodo,
                formato="pdf",
                confianca=0.95,
            )

        if "SANTANDER" in texto_upper or "4220 XXXX XXXX 7342" in texto_completo:
            periodo = _extrair_periodo_santander(texto_completo)
            return DeteccaoArquivo(
                banco="santander",
                tipo="cartao",
                pessoa="andre",
                subtipo="",
                periodo=periodo,
                formato="pdf",
                confianca=0.95,
            )

    except Exception as erro:
        logger.warning("Erro ao analisar PDF %s: %s", caminho, erro)
    finally:
        pdf.close()

    return None


def detectar_arquivo(caminho: Path) -> Optional[DeteccaoArquivo]:
    """Detecta banco, tipo, pessoa e período de um arquivo financeiro pelo conteúdo.

    Args:
        caminho: Caminho para o arquivo a ser analisado.

    Returns:
        DeteccaoArquivo com os dados detectados, ou None se não reconhecido.
    """
    if not caminho.exists():
        logger.warning("Arquivo não encontrado: %s", caminho)
        return None

    if not caminho.is_file():
        logger.debug("Caminho não é arquivo: %s", caminho)
        return None

    sufixo = caminho.suffix.lower()

    detectores: dict[str, type] = {
        ".csv": _detectar_csv,
        ".xlsx": _detectar_xlsx,
        ".xls": _detectar_xls,
        ".pdf": _detectar_pdf,
    }

    detector = detectores.get(sufixo)
    if detector is None:
        logger.debug("Formato não suportado: %s", sufixo)
        return None

    logger.info("Analisando arquivo: %s", caminho.name)
    resultado = detector(caminho)

    if resultado:
        logger.info(
            "Detectado: banco=%s tipo=%s pessoa=%s subtipo=%s período=%s confiança=%.2f",
            resultado.banco,
            resultado.tipo,
            resultado.pessoa,
            resultado.subtipo,
            resultado.periodo,
            resultado.confianca,
        )
    else:
        logger.warning("Não foi possível identificar: %s", caminho.name)

    return resultado


# "Aquele que tem um porquê para viver pode suportar quase qualquer como." -- Friedrich Nietzsche
