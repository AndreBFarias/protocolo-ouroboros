"""Orquestrador principal do pipeline ETL financeiro."""

import argparse
import hashlib
import os
import re
import shutil
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from src.extractors.contracheque_pdf import processar_holerites
from src.load.relatorio import gerar_relatorios
from src.load.xlsx_writer import gerar_xlsx
from src.transform.canonicalizer_casal import (
    e_transferencia_do_casal,
    variantes_curtas,
)
from src.transform.categorizer import Categorizer
from src.transform.deduplicator import deduplicar
from src.transform.irpf_tagger import aplicar_tags_irpf
from src.transform.normalizer import normalizar_transacao
from src.utils.lockfile import Lockfile, LockfileOcupado
from src.utils.logger import configurar_logger

logger = configurar_logger("pipeline")

RAIZ = Path(__file__).parent.parent
DIR_RAW = RAIZ / "data" / "raw"
DIR_OUTPUT = RAIZ / "data" / "output"
DIR_HISTORICO = RAIZ / "data" / "historico"
CONTROLE_ANTIGO = DIR_HISTORICO / "controle_antigo.xlsx"
PATH_GRAFO = DIR_OUTPUT / "grafo.sqlite"
DIR_BACKUP_GRAFO = DIR_OUTPUT / "backup"
PREFIXO_BACKUP_GRAFO = "grafo_"
RETENCAO_DIAS_RECENTES = 7
RETENCAO_SEMANAS_ADICIONAIS = 4

# Sprint INFRA-CONCORRENCIA-PIDFILE (2026-05-16): lock canônico do pipeline.
# Dashboard consome o mesmo path para detectar pipeline ativo e exibir toast.
PATH_LOCKFILE = RAIZ / "data" / ".pipeline.lock"


# Sprint INFRA-DESCOBRIR-EXTRATORES-REFATORA (2026-05-17): lista declarativa
# canônica substitui ~140 linhas de try/except idênticos. Ordem importa
# (pipeline serial), por isso lista explícita em vez de pkgutil.walk.
# Suporte a desabilitar via env: OUROBOROS_EXTRATORES_DESABILITADOS=nubank_cartao,c6_cc
EXTRATORES_CANONICOS: list[tuple[str, str]] = [
    # (módulo, nome_classe) — ordem é a mesma de antes da refatora
    ("src.extractors.nubank_cartao", "ExtratorNubankCartao"),
    ("src.extractors.nubank_cc", "ExtratorNubankCC"),
    ("src.extractors.c6_cc", "ExtratorC6CC"),
    ("src.extractors.c6_cartao", "ExtratorC6Cartao"),
    ("src.extractors.itau_pdf", "ExtratorItauPDF"),
    ("src.extractors.santander_pdf", "ExtratorSantanderPDF"),
    ("src.extractors.energia_ocr", "ExtratorEnergiaOCR"),
    ("src.extractors.ofx_parser", "ExtratorOFX"),
    ("src.extractors.cupom_garantia_estendida_pdf", "ExtratorCupomGarantiaEstendida"),
    ("src.extractors.nfce_pdf", "ExtratorNfcePDF"),
    ("src.extractors.danfe_pdf", "ExtratorDanfePDF"),
    ("src.extractors.cupom_termico_foto", "ExtratorCupomTermicoFoto"),
    ("src.extractors.comprovante_pix_foto", "ExtratorComprovantePixFoto"),
    ("src.extractors.xml_nfe", "ExtratorXmlNFe"),
    ("src.extractors.receita_medica", "ExtratorReceitaMedica"),
    ("src.extractors.garantia", "ExtratorGarantiaFabricante"),
    ("src.extractors.das_parcsn_pdf", "ExtratorDASPARCSNPDF"),
    ("src.extractors.dirpf_dec", "ExtratorDIRPFDec"),
    ("src.extractors.boleto_pdf", "ExtratorBoletoPDF"),
    ("src.extractors.recibo_nao_fiscal", "ExtratorReciboNaoFiscal"),
]


def _descobrir_extratores(desabilitados: set[str] | None = None) -> list:
    """Importa e retorna classes de extratores canônicas.

    Sprint INFRA-DESCOBRIR-EXTRATORES-REFATORA (2026-05-17): substitui
    ~140 linhas de try/except por iteração sobre `EXTRATORES_CANONICOS`.
    Mantém falha-soft: extrator não-importável é apenas logado.

    Args:
        desabilitados: nomes curtos de módulos (ex: ``{"nubank_cartao"}``)
            a pular. Se ``None``, lê de ``OUROBOROS_EXTRATORES_DESABILITADOS``
            (CSV em env var). Útil para debug isolado de outros extratores.

    Returns:
        Lista de classes de extrator na ordem canônica (pipeline depende dela).
    """
    import importlib

    if desabilitados is None:
        env_val = os.environ.get("OUROBOROS_EXTRATORES_DESABILITADOS", "")
        desabilitados = {x.strip() for x in env_val.split(",") if x.strip()}

    extratores = []
    for modulo, classname in EXTRATORES_CANONICOS:
        nome_curto = modulo.rsplit(".", 1)[-1]
        if nome_curto in desabilitados:
            logger.info("Extrator %s desabilitado via env.", nome_curto)
            continue
        try:
            mod = importlib.import_module(modulo)
            cls = getattr(mod, classname)
            extratores.append(cls)
        except (ImportError, AttributeError) as e:
            logger.warning("Extrator %s indisponível: %s", nome_curto, e)
    return extratores


def _escanear_arquivos(diretorio: Path) -> list[Path]:
    """Escaneia recursivamente todos os arquivos em data/raw/."""
    extensoes = {
        ".csv",
        ".xlsx",
        ".xls",
        ".xml",
        ".pdf",
        ".ofx",
        ".jpg",
        ".jpeg",
        ".png",
        ".heic",
        ".heif",
        ".dec",  # P3.1 2026-04-23: DIRPF .DEC (Receita Federal)
    }
    arquivos = []

    for arquivo in diretorio.rglob("*"):
        if arquivo.is_file() and arquivo.suffix.lower() in extensoes:
            # Ignorar arquivos duplicados com sufixo (1), (2)
            if " (1)" in arquivo.stem or " (2)" in arquivo.stem:
                logger.debug("Ignorando duplicata de download: %s", arquivo.name)
                continue
            arquivos.append(arquivo)

    logger.info("Encontrados %d arquivos para processar em %s", len(arquivos), diretorio)
    return sorted(arquivos)


def _extrair_tudo(arquivos: list[Path], classes_extratores: list) -> list[dict]:
    """Executa extração de todos os arquivos com os extratores disponíveis."""
    transacoes_brutas: list[dict] = []
    arquivos_processados = 0
    arquivos_ignorados = 0

    for arquivo in arquivos:
        processado = False
        for cls_extrator in classes_extratores:
            try:
                extrator = cls_extrator(arquivo)
                if extrator.pode_processar(arquivo):
                    resultado = extrator.extrair()
                    for t in resultado:
                        transacao_norm = normalizar_transacao(
                            data_transacao=t.data,
                            valor=t.valor,
                            descricao=t.descricao,
                            banco_origem=t.banco_origem,
                            tipo_extrato="cartao"
                            if "cartao" in t.banco_origem.lower() or t.forma_pagamento == "Crédito"
                            else "cc",
                            identificador=t.identificador,
                            subtipo=_inferir_subtipo(arquivo),
                            arquivo_origem=str(arquivo),
                            tipo_sugerido=t.tipo,
                            valor_original_com_sinal=t.valor,
                            virtual=getattr(t, "_virtual", False),
                        )
                        transacoes_brutas.append(transacao_norm)

                    arquivos_processados += 1
                    logger.info(
                        "Extraídas %d transações de %s (%s)",
                        len(resultado),
                        arquivo.name,
                        cls_extrator.__name__,
                    )
                    processado = True
                    break
            except Exception as e:
                logger.error(
                    "Erro ao processar %s com %s: %s", arquivo.name, cls_extrator.__name__, e
                )

        if not processado:
            arquivos_ignorados += 1
            logger.warning("Nenhum extrator compatível para: %s", arquivo.name)

    logger.info(
        "Extração concluída: %d arquivos processados, %d ignorados, %d transações brutas",
        arquivos_processados,
        arquivos_ignorados,
        len(transacoes_brutas),
    )
    return transacoes_brutas


def _inferir_subtipo(arquivo: Path) -> str | None:
    """Infere subtipo (pf/pj) pelo caminho do arquivo."""
    partes = str(arquivo).lower()
    if "pj" in partes:
        return "pj"
    if "pf" in partes:
        return "pf"
    return None


def _importar_historico() -> list[dict]:
    """Importa transações do XLSX histórico (ago/2022 - jul/2023)."""
    if not CONTROLE_ANTIGO.exists():
        logger.info("Arquivo histórico não encontrado: %s", CONTROLE_ANTIGO)
        return []

    import openpyxl

    transacoes = []
    try:
        wb = openpyxl.load_workbook(CONTROLE_ANTIGO)
        if "Extrato" not in wb.sheetnames:
            logger.warning("Aba 'Extrato' não encontrada no histórico")
            return []

        ws = wb["Extrato"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            data_val = row[0] if len(row) > 0 else None
            gasto = row[1] if len(row) > 1 else None
            forma = row[2] if len(row) > 2 else ""
            local = row[3] if len(row) > 3 else ""
            quem = row[4] if len(row) > 4 else ""
            categoria = row[5] if len(row) > 5 else None
            classificacao = row[6] if len(row) > 6 else None

            if data_val is None or gasto is None:
                continue

            if isinstance(data_val, datetime):
                data_date = data_val.date()
            elif isinstance(data_val, str):
                try:
                    data_date = datetime.strptime(data_val, "%Y-%m-%d").date()
                except ValueError:
                    continue
            else:
                continue

            # Normalizar classificação corrompida do histórico
            clf_raw = str(classificacao).strip() if classificacao else "Questionável"
            clf_map = {
                "Obrigatório": "Obrigatório",
                "Obrigatórios": "Obrigatório",
                "Questionável": "Questionável",
                "Supérfluo": "Supérfluo",
            }
            clf_normalizada = clf_map.get(clf_raw, "Questionável")

            transacoes.append(
                {
                    "data": data_date,
                    "valor": abs(float(gasto)) if isinstance(gasto, (int, float)) else 0,
                    "forma_pagamento": str(forma) if forma else "Débito",
                    "local": str(local) if local else "",
                    "quem": str(quem) if quem else "Casal",
                    "categoria": str(categoria) if categoria else "Outros",
                    "classificacao": clf_normalizada,
                    "banco_origem": "Histórico",
                    "tipo": "Despesa",
                    "mes_ref": data_date.strftime("%Y-%m"),
                    "tag_irpf": None,
                    "obs": "Importado do histórico",
                    "_identificador": f"hist_{data_date.isoformat()}_{gasto}_{local}",
                    "_descricao_original": str(local),
                    "_arquivo_origem": str(CONTROLE_ANTIGO),
                }
            )

        logger.info("Histórico importado: %d transações (ago/2022 - jul/2023)", len(transacoes))
    except Exception as e:
        logger.error("Erro ao importar histórico: %s", e)

    return transacoes


def _reclassificar_ti_orfas(transacoes: list[dict]) -> list[dict]:
    """Reclassifica TIs cuja descrição NÃO bate identidade do casal.

    Sprint 68b: rede de segurança pós-deduplicação. Qualquer transação
    marcada como `Transferência Interna` cuja descrição original não
    passa pelo matcher formal `e_transferencia_do_casal` é degradada
    para `Despesa` (quando valor negativo) ou `Receita` (valor positivo).

    Exceções operacionais legítimas (pagamento de fatura do próprio
    banco, resgate/aplicação CDB/RDB, agência 6450 Itaú) são preservadas
    via heurísticas textuais simples -- essas regras ainda vivem nos
    extratores e já foram aplicadas antes deste ponto; aqui o objetivo é
    apenas caçar falsos-positivos que escaparam (ex: linhas importadas
    do `controle_antigo.xlsx` cujo `local` cita terceiro homônimo).
    """
    regex_operacional = re.compile(
        r"PAGAMENTO\s+DE\s+FATURA|PGTO\s+FAT\s+CARTAO|PGTO\s+FATURA|"
        r"Fatura\s+de\s+cart[aã]o|PAGAMENTO\s+FATURA\s+NU|PAGTO\s+NU\s*PAGAMENT|"
        r"DEBITO\s+DE\s+CARTAO|DEB\s+CART|PIX\s+QRS\s+BANCO\s+SANTA|"
        r"CDB\s+C6|LIM\.\s*GARANT|RESGATE\s+CDB|APLICA[CÇ][AÃ]O\s+CDB|"
        r"AG\s*6450|AGENCIA\s+6450|Valor\s+adicionado\s+na\s+conta|Pix\s+no\s+Cr[eé]dito",
        re.IGNORECASE,
    )

    # Sprint AUDIT-TI-RECLASSIFICA-RASTREAMENTO (2026-05-17): rastreamento
    # granular para auditoria. Cada transação reclassificada ganha flag
    # `_reclassificada_68b=True` + `_tipo_anterior` + amostras gravadas em
    # log estruturado data/output/reclassificacao_ti_orfas_<ts>.json.
    reclassificadas = 0
    amostras_log: list[dict] = []
    for t in transacoes:
        if t.get("tipo") != "Transferência Interna":
            continue

        # Sprint 82b: espelho virtual de cartão é TI por design (contraparte
        # de pagamento de fatura). Não depende de match textual do casal
        # nem do regex operacional -- preserva a flag e pula a degradação.
        if t.get("_virtual"):
            continue

        descricao = str(t.get("_descricao_original") or t.get("local") or "")

        if e_transferencia_do_casal(descricao):
            continue

        if regex_operacional.search(descricao):
            continue

        try:
            valor = float(t.get("valor", 0) or 0)
        except (TypeError, ValueError):
            valor = 0.0

        novo_tipo = "Receita" if valor > 0 else "Despesa"
        t["_tipo_anterior"] = "Transferência Interna"
        t["_reclassificada_68b"] = True
        t["_razao_reclassificacao"] = "nao_bate_casal_nem_regex_operacional"
        t["tipo"] = novo_tipo
        reclassificadas += 1

        # Amostra para log (max 100 para não explodir JSON):
        if len(amostras_log) < 100:
            amostras_log.append(
                {
                    "data": (
                        t["data"].isoformat()
                        if hasattr(t.get("data"), "isoformat")
                        else str(t.get("data", ""))
                    ),
                    "valor": valor,
                    "local": descricao[:80],
                    "banco_origem": t.get("banco_origem", ""),
                    "tipo_anterior": "Transferência Interna",
                    "tipo_novo": novo_tipo,
                    "razao": "nao_bate_casal_nem_regex_operacional",
                }
            )

    if reclassificadas:
        logger.info(
            "Sprint 68b: %d TI órfãs reclassificadas (sem match casal + sem regra operacional)",
            reclassificadas,
        )
        _gravar_log_reclassificacao_ti_orfas(reclassificadas, amostras_log)
    return transacoes


def _gravar_log_reclassificacao_ti_orfas(total: int, amostras: list[dict]) -> Path | None:
    """Grava log estruturado de reclassificação TI órfã.

    Sprint AUDIT-TI-RECLASSIFICA-RASTREAMENTO (2026-05-17): rastreabilidade
    para auditoria humana. Arquivo gitignored (sob data/output/).
    """
    import json

    ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    destino = DIR_OUTPUT / f"reclassificacao_ti_orfas_{ts}.json"
    payload = {
        "executado_em": datetime.now().isoformat(),
        "total_revertidas": total,
        "amostras": amostras,
        "_doc": (
            "Transferências Internas que NÃO bateram nem e_transferencia_do_casal "
            "nem regex_operacional foram degradadas para Despesa/Receita. Revisar "
            "amostras para confirmar que não há transferência legítima sendo "
            "degradada falsamente."
        ),
    }
    try:
        destino.parent.mkdir(parents=True, exist_ok=True)
        destino.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )
        return destino
    except OSError as exc:
        logger.warning("Falha ao gravar log reclassificacao_ti_orfas: %s", exc)
        return None


def _promover_variantes_para_ti(transacoes: list[dict]) -> list[dict]:
    """Promove Receita/Despesa para Transferência Interna via variantes curtas.

    Sprint 82: rede de captura pós-reclassificação. Qualquer transação
    ainda marcada como Receita ou Despesa cuja descrição casa o nível 2
    do matcher (`variantes_curtas(descricao, banco_origem)`) é promovida  # noqa: accent
    a Transferência Interna. Preserva a simetria com
    `_reclassificar_ti_orfas`: a função anterior degrada falsos-positivos;
    esta captura os falsos-negativos do matcher rigoroso.

    Não toca linhas já marcadas como Transferência Interna (evita
    dupla-contagem) nem Imposto.
    """
    promovidas = 0
    for t in transacoes:
        tipo_atual = t.get("tipo")
        if tipo_atual not in {"Receita", "Despesa"}:
            continue

        descricao = str(t.get("_descricao_original") or t.get("local") or "")
        banco = str(t.get("banco_origem") or "")
        if not descricao or not banco:
            continue

        if e_transferencia_do_casal(descricao):
            # Casou no matcher rigoroso mas tipo não é TI -- aproveita
            # e reclassifica aqui (defesa em profundidade).
            t["tipo"] = "Transferência Interna"
            promovidas += 1
            continue

        if variantes_curtas(descricao, banco):
            t["tipo"] = "Transferência Interna"
            promovidas += 1

    if promovidas:
        logger.info(
            "Sprint 82: %d transações promovidas a Transferência Interna via variantes curtas",
            promovidas,
        )
    return transacoes


def _filtrar_por_mes(transacoes: list[dict], mes: str) -> list[dict]:
    """Filtra transações por mês específico."""
    return [t for t in transacoes if t.get("mes_ref") == mes]


def _executar_linking_documentos() -> None:
    """Aciona o motor de linking documento->transação quando o grafo existe.

    Importação lazy para não pagar custo quando o grafo não existe ainda
    (usuário que roda só o pipeline XLSX clássico). Falhas são logadas
    como warning e não quebram a execução.
    """
    try:
        from src.graph.db import caminho_padrao
    except ImportError as erro:
        logger.warning("Módulo src.graph indisponível: %s -- linking ignorado", erro)
        return

    caminho_grafo = caminho_padrao()
    if not caminho_grafo.exists():
        logger.info(
            "Grafo SQLite ausente em %s -- linking de documentos pulado",
            caminho_grafo,
        )
        return

    try:
        from src.graph.db import GrafoDB
        from src.graph.linking import (
            linkar_documentos_a_transacoes,
            linkar_pix_transacao,
        )

        # Sprint INFRA-PIPELINE-TRANSACIONALIDADE: rollback granular se
        # qualquer função interna crashar mid-linking; estágios anteriores
        # do pipeline ficam preservados no grafo.
        with GrafoDB(caminho_grafo) as db, db.transaction():
            stats = linkar_documentos_a_transacoes(db)
            stats_pix = linkar_pix_transacao(db)
        logger.info("Linking documento->transação: %s", stats)
        logger.info("Linking PIX dedicado: %s", stats_pix)
    except Exception as erro:
        _registrar_falha_pipeline_estruturada("linking_documentos", erro)
        logger.warning("Linking de documentos falhou: %s", erro)


def _executar_er_produtos() -> None:
    """Aciona o entity resolution de produtos quando o grafo existe.

    Roda DEPOIS do linking de documentos (Sprint 48) para que nodes `item` já
    estejam em posição estável. Importação lazy e falhas não-fatais (mesma
    política do linking).
    """
    try:
        from src.graph.db import caminho_padrao
    except ImportError as erro:
        logger.warning("Módulo src.graph indisponível: %s -- ER produtos ignorado", erro)
        return

    caminho_grafo = caminho_padrao()
    if not caminho_grafo.exists():
        logger.info(
            "Grafo SQLite ausente em %s -- ER de produtos pulado",
            caminho_grafo,
        )
        return

    try:
        from src.graph.db import GrafoDB
        from src.graph.er_produtos import executar_er_produtos

        # Transação envolve toda a fase de ER: se um item crashar no meio,
        # produtos canônicos parciais NÃO ficam no grafo.
        with GrafoDB(caminho_grafo) as db, db.transaction():
            stats = executar_er_produtos(db)
        logger.info("ER de produtos: %s", stats)
    except Exception as erro:
        _registrar_falha_pipeline_estruturada("er_produtos", erro)
        logger.warning("ER de produtos falhou: %s", erro)


def _executar_dossie_snapshot() -> None:
    """Atualiza ``data/output/graduacao_tipos.json`` a partir dos dossies.

    Integrado em 2026-05-13: cada `./run.sh --tudo` regenera o snapshot global
    de graduacao por tipo documental, consumido pelo dashboard. Falha-soft.
    Ver `docs/CICLO_GRADUACAO_OPERACIONAL.md` para o ritual completo.
    """
    try:
        from scripts.dossie_tipo import cmd_snapshot
    except ImportError as erro:
        logger.warning("dossie_tipo indisponivel: %s", erro)
        return
    try:
        cmd_snapshot()
    except Exception as erro:
        logger.warning("snapshot de graduacao falhou: %s", erro)


def _executar_skill_d7_log() -> None:
    """Gera ``data/output/skill_d7_log.json`` (Sprint INFRA-SKILLS-D7-LOG).

    Snapshot estruturado do classificador D7, consumido pela página
    ``src/dashboard/paginas/skills_d7.py``. Integração feita em 2026-05-13
    após auditoria detectar que o script existia em `scripts/gerar_skill_d7_log.py`
    mas não era invocado automaticamente. Falha-soft.
    """
    try:
        from scripts.gerar_skill_d7_log import gerar_snapshot
    except ImportError as erro:
        logger.warning("Snapshot D7 indisponível (import falhou): %s", erro)
        return

    try:
        import json

        snapshot = gerar_snapshot()
        caminho_out = DIR_OUTPUT / "skill_d7_log.json"
        caminho_out.write_text(
            json.dumps(snapshot, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        logger.info("Snapshot D7 gravado: %s", caminho_out)
    except Exception as erro:
        logger.warning("Geração do snapshot D7 falhou: %s", erro)


def _executar_item_categorizer() -> None:
    """Aplica categorização por regex a todos os nodes `item` do grafo.

    Sprint 50: roda DEPOIS do ER de produtos (Sprint 49) para que a agregação
    por `produto_canonico` já exista antes da atribuição de categoria. Também
    importação lazy e falhas não-fatais.
    """
    try:
        from src.graph.db import caminho_padrao
    except ImportError as erro:
        logger.warning(
            "Módulo src.graph indisponível: %s -- categorização de itens pulada",
            erro,
        )
        return

    caminho_grafo = caminho_padrao()
    if not caminho_grafo.exists():
        logger.info(
            "Grafo SQLite ausente em %s -- categorização de itens pulada",
            caminho_grafo,
        )
        return

    try:
        from src.graph.db import GrafoDB
        from src.transform.item_categorizer import categorizar_todos_items_no_grafo

        # Transação evita categorização parcial de itens (rollback se crash
        # antes de processar todos os nodes `item`).
        with GrafoDB(caminho_grafo) as db, db.transaction():
            stats = categorizar_todos_items_no_grafo(db)
        logger.info("Categorização de itens: %s", stats)
    except Exception as erro:
        _registrar_falha_pipeline_estruturada("item_categorizer", erro)
        logger.warning("Categorização de itens falhou: %s", erro)


# ---------------------------------------------------------------------------
# Backup automatico do grafo (Sprint INFRA-BACKUP-GRAFO-AUTOMATIZADO)  # noqa: accent
# ---------------------------------------------------------------------------


def _sha256_arquivo(caminho: Path) -> str:
    """Calcula SHA-256 hex do conteúdo de um arquivo em chunks."""  # noqa: accent
    h = hashlib.sha256()
    with caminho.open("rb") as fp:
        for chunk in iter(lambda: fp.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _ts_to_timestamp(ts: str) -> datetime | None:
    """Parse seguro do timestamp `YYYY-MM-DD_HHMMSS` no nome do backup."""
    try:
        return datetime.strptime(ts, "%Y-%m-%d_%H%M%S")
    except ValueError:
        return None


def _executar_backup_grafo(
    grafo: Path = PATH_GRAFO, dir_backup: Path = DIR_BACKUP_GRAFO
) -> Path | None:
    """Snapshot pre-pipeline do grafo SQLite + sha256.

    Retorna o caminho do backup criado, ou None se o grafo não existir
    (pipeline em estado inicial -- nada a preservar).
    """
    if not grafo.exists():
        logger.info("Backup grafo: arquivo origem ausente em %s; skip.", grafo)
        return None
    dir_backup.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    destino = dir_backup / f"{PREFIXO_BACKUP_GRAFO}{ts}.sqlite"
    shutil.copy2(grafo, destino)
    sha_path = destino.with_suffix(destino.suffix + ".sha256")
    sha = _sha256_arquivo(destino)
    sha_path.write_text(f"{sha}  {destino.name}\n", encoding="utf-8")
    logger.info("Backup grafo: %s (sha256=%s...)", destino.name, sha[:12])
    _aplicar_retencao_backups_grafo(dir_backup)
    return destino


def _aplicar_retencao_backups_grafo(dir_backup: Path) -> list[Path]:
    """Aplica politica: ultimos 7 dias completos + 1 por semana das 4 anteriores.

    Retorna lista de arquivos deletados.
    """
    if not dir_backup.exists():
        return []
    candidatos: list[tuple[datetime, Path]] = []
    for p in dir_backup.glob(f"{PREFIXO_BACKUP_GRAFO}*.sqlite"):
        ts_str = p.stem[len(PREFIXO_BACKUP_GRAFO) :]
        ts = _ts_to_timestamp(ts_str)
        if ts is None:
            continue
        candidatos.append((ts, p))
    candidatos.sort(reverse=True)

    a_manter: set[Path] = set()
    agora = datetime.now()
    limite_recente = agora - timedelta(days=RETENCAO_DIAS_RECENTES)

    # Camada 1: tudo dentro dos ultimos 7 dias
    for ts, p in candidatos:
        if ts >= limite_recente:
            a_manter.add(p)

    # Camada 2: 1 backup por semana para as 4 semanas anteriores aos 7 dias
    semanas_cobertas: set[int] = set()
    for ts, p in candidatos:
        if ts >= limite_recente:
            continue
        # Número da semana relativa: 0=primeira semana antes do limite, 1=segunda, etc.
        dias_alem = (limite_recente - ts).days
        idx_semana = dias_alem // 7
        if idx_semana >= RETENCAO_SEMANAS_ADICIONAIS:
            continue
        if idx_semana not in semanas_cobertas:
            a_manter.add(p)
            semanas_cobertas.add(idx_semana)

    deletados: list[Path] = []
    for _ts, p in candidatos:
        if p not in a_manter:
            sha_path = p.with_suffix(p.suffix + ".sha256")
            p.unlink(missing_ok=True)
            sha_path.unlink(missing_ok=True)
            deletados.append(p)
    if deletados:
        logger.info("Backup grafo: retencao removeu %d arquivos antigos.", len(deletados))
    return deletados


def _restaurar_grafo_de_backup(
    timestamp: str, grafo: Path = PATH_GRAFO, dir_backup: Path = DIR_BACKUP_GRAFO
) -> int:
    """Restaura grafo a partir de backup identificado pelo timestamp.

    Valida checksum antes de sobrescrever. Retorna 0 em sucesso, 1 em falha.
    """
    backup = dir_backup / f"{PREFIXO_BACKUP_GRAFO}{timestamp}.sqlite"
    sha_path = backup.with_suffix(backup.suffix + ".sha256")
    if not backup.exists():
        logger.error("Backup ausente: %s", backup)
        return 1
    if not sha_path.exists():
        logger.error("Checksum ausente: %s", sha_path)
        return 1
    sha_gravado = sha_path.read_text(encoding="utf-8").split()[0].strip()
    if len(sha_gravado) != 64:
        logger.error(
            "Checksum malformado: .sha256 tem %d chars (esperado 64). "
            "Provável write parcial. Regenere com `sha256sum %s > %s`.",
            len(sha_gravado),
            backup,
            sha_path,
        )
        return 1
    if not re.match(r"^[0-9a-f]{64}$", sha_gravado):
        logger.error(
            "Checksum corrompido: caracteres inválidos. "
            "Arquivo .sha256 pode estar comprometido. Tente backup anterior."
        )
        return 1
    sha_atual = _sha256_arquivo(backup)
    if sha_gravado != sha_atual:
        logger.error(
            "Conteúdo do backup corrompido: sha gravado=%s vs sha calculado=%s. "
            "Backup inutilizável. Tente versão anterior.",
            sha_gravado[:12],
            sha_atual[:12],
        )
        return 1
    if grafo.exists():
        ts_recuo = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        recuo = grafo.with_suffix(grafo.suffix + f".pre_restore_{ts_recuo}")
        shutil.copy2(grafo, recuo)
        logger.info("Grafo atual preservado em %s antes do restore.", recuo.name)
    shutil.copy2(backup, grafo)
    logger.info("Grafo restaurado de %s.", backup.name)
    return 0


def _registrar_falha_pipeline_estruturada(estagio: str, erro: Exception) -> Path | None:
    """Grava ``logs/pipeline_falha_<ts>.json`` com diagnóstico estruturado.

    Sprint INFRA-PIPELINE-TRANSACIONALIDADE (2026-05-15): quando um estágio
    do pipeline crasha, registramos: estágio canônico, traceback completo,
    timestamp ISO, e ponteiro para o backup automático mais recente do
    grafo (sprint INFRA-BACKUP-GRAFO-AUTOMATIZADO). Sprint encadeada
    INFRA-PIPELINE-TX-RESTORE-AUTOMATICO consome este JSON para decidir
    rollback global.

    Falha-soft: erro de escrita do log NÃO propaga (não queremos que o
    log de falha mate o pipeline duas vezes).
    """
    try:
        import json as _json
        import traceback

        ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        log_dir = RAIZ / "logs"
        log_dir.mkdir(parents=True, exist_ok=True)
        destino = log_dir / f"pipeline_falha_{ts}.json"
        ultimo_backup: str | None = None
        backups = sorted(DIR_BACKUP_GRAFO.glob(f"{PREFIXO_BACKUP_GRAFO}*.sqlite"))
        if backups:
            ultimo_backup = backups[-1].name
        registro = {
            "ts": datetime.now().isoformat(),
            "estagio": estagio,
            "erro_tipo": type(erro).__name__,
            "erro_mensagem": str(erro),
            "traceback": traceback.format_exc(),
            "ultimo_backup_grafo": ultimo_backup,
        }
        destino.write_text(_json.dumps(registro, ensure_ascii=False, indent=2), encoding="utf-8")
        logger.info("Falha de pipeline registrada em %s", destino.name)
        return destino
    except Exception as registrar_erro:
        logger.warning(
            "Falhou ao registrar log estruturado da falha de pipeline: %s",
            registrar_erro,
        )
        return None


_ESTAGIO_ATUAL: str = "init"


def executar(mes: str | None = None, processar_tudo: bool = False) -> None:
    """Executa o pipeline completo.

    Sprint INFRA-PIPELINE-TX-RESTORE-AUTOMATICO (2026-05-15): após o backup
    pré-pipeline, o corpo inteiro roda dentro de `try/except`. Qualquer
    exceção não-tratada por estágios internos dispara:

    1. Log estruturado em `logs/pipeline_falha_<ts>.json` (com estágio atual).
    2. Tentativa de restore automático do backup pré-execução (se existe).
    3. Re-raise da exceção original (falha não fica silenciosa).

    Estágios mutadores do grafo já têm rollback granular via
    `with db.transaction():` (sprint INFRA-PIPELINE-TRANSACIONALIDADE).
    Este restore é a defesa em camadas externa para falhas catastróficas
    (OOM kill, signal interrupt, corrupção SQLite).
    """
    global _ESTAGIO_ATUAL
    logger.info("=== Protocolo Ouroboros -- Pipeline ===")

    # Sprint INFRA-CONCORRENCIA-PIDFILE (2026-05-16): lock exclusivo serializa
    # escritas concorrentes ao grafo + XLSX. Defesa em camadas com flock no
    # run.sh (padrao (n)). Falha-fast com mensagem amigavel se outro pipeline
    # esta rodando.
    descricao = f"pipeline mes={mes}" if mes else "pipeline tudo" if processar_tudo else "pipeline"
    try:
        ctx_lock = Lockfile(PATH_LOCKFILE, descricao)
        ctx_lock.__enter__()
    except LockfileOcupado as exc:
        logger.error(
            "Pipeline abortado: outra instancia esta rodando (PID=%s). "
            "Aguarde ou mate o processo dono do lock %s.",
            exc.pid_dono,
            exc.path,
        )
        sys.exit(2)

    try:
        # 0. Backup automatico do grafo (Sprint INFRA-BACKUP-GRAFO-AUTOMATIZADO).
        # Padrao (m) branch reversivel: snapshot pre-execucao permite rollback  # noqa: accent
        # se pipeline crashar mid-ETL.
        _ESTAGIO_ATUAL = "backup_grafo"
        backup_destino = _executar_backup_grafo()
        backup_ts: str | None = None
        if backup_destino is not None:
            backup_ts = backup_destino.stem[len(PREFIXO_BACKUP_GRAFO) :]

        _executar_corpo_pipeline(mes, processar_tudo)
    except Exception as exc:
        _registrar_falha_pipeline_estruturada(_ESTAGIO_ATUAL, exc)
        if backup_ts is not None:
            logger.error(
                "Pipeline crashou no estágio %s. Tentando restore automático do backup %s.",
                _ESTAGIO_ATUAL,
                backup_ts,
            )
            try:
                _restaurar_grafo_de_backup(backup_ts)
            except Exception as restore_erro:
                logger.error(
                    "Restore automático falhou: %s. Backup preservado em %s.",
                    restore_erro,
                    DIR_BACKUP_GRAFO,
                )
        else:
            logger.error(
                "Pipeline crashou no estágio %s sem backup pré-execução "
                "(primeira run). Restore não aplicável.",
                _ESTAGIO_ATUAL,
            )
        raise
    finally:
        # Libera lockfile mesmo em caso de exceção. Padrao (m) reversivel.
        ctx_lock.__exit__(None, None, None)


# ---------------------------------------------------------------------------
# Sprint INFRA-PIPELINE-FASES-ISOLADAS (2026-05-17): split do corpo do pipeline
# em 17 fases isoladas. Cada fase recebe um `ctx: dict` mutável compartilhado
# e devolve um dict de `stats` para o log estruturado. Ordem é preservada
# bit-a-bit em relação à versão monolítica (pipeline serial depende disso).
#
# Contrato de uma fase: `fase_xxx(ctx: dict) -> dict[str, Any]`.
# Chaves esperadas em `ctx` ao longo da execução:
#   - mes: str | None
#   - processar_tudo: bool
#   - classes_extratores: list (após fase 1)
#   - arquivos: list[Path] (após fase 2)
#   - transacoes: list[dict] (mutada do passo 3 em diante)  # noqa: accent
#   - transacoes_filtradas: list[dict] (após fase 8)  # noqa: accent
#   - contracheques: list[dict] (após fase 10)
#   - caminho_xlsx: Path (após fase 11)
# ---------------------------------------------------------------------------


def fase_descobrir_extratores(ctx: dict[str, Any]) -> dict[str, Any]:
    """Fase 1: descobre classes de extratores disponíveis."""
    global _ESTAGIO_ATUAL
    _ESTAGIO_ATUAL = "descobrir_extratores"
    classes_extratores = _descobrir_extratores()
    logger.info("Extratores disponíveis: %d", len(classes_extratores))
    ctx["classes_extratores"] = classes_extratores
    return {"n_extratores": len(classes_extratores)}


def fase_escanear_arquivos(ctx: dict[str, Any]) -> dict[str, Any]:
    """Fase 2: escaneia DIR_RAW por arquivos elegíveis."""
    arquivos = _escanear_arquivos(DIR_RAW)
    ctx["arquivos"] = arquivos
    return {"n_arquivos": len(arquivos)}


def fase_extrair(ctx: dict[str, Any]) -> dict[str, Any]:
    """Fase 3: aplica extratores a cada arquivo, produzindo transações brutas."""
    transacoes = _extrair_tudo(ctx["arquivos"], ctx["classes_extratores"])
    ctx["transacoes"] = transacoes
    return {"n_extraidas": len(transacoes)}


def fase_importar_historico(ctx: dict[str, Any]) -> dict[str, Any]:
    """Fase 4: anexa transações do histórico legado (controle_antigo)."""
    historico = _importar_historico()
    ctx["transacoes"].extend(historico)
    return {"n_historico": len(historico), "n_apos_historico": len(ctx["transacoes"])}


def fase_deduplicar(ctx: dict[str, Any]) -> dict[str, Any]:
    """Fase 5: remove duplicatas via hash_fuzzy (canonicalização + dedup pass 2b)."""
    antes = len(ctx["transacoes"])
    ctx["transacoes"] = deduplicar(ctx["transacoes"])
    depois = len(ctx["transacoes"])
    return {"n_antes": antes, "n_depois": depois, "n_removidas": antes - depois}


def fase_categorizar(ctx: dict[str, Any]) -> dict[str, Any]:
    """Fase 6: aplica regras de categoria + reclassificação TI órfã + promoção
    de variantes curtas. Sub-estágios 6/6b/6c agrupados porque devem rodar em
    sequência rígida (ordem comentada na versão monolítica)."""
    categorizer = Categorizer()
    ctx["transacoes"] = categorizer.categorizar_lote(ctx["transacoes"])
    # 6b. Reclassificar TIs órfãs (Sprint 68b) -- rede de segurança
    # pós-categorização contra falsos-positivos residuais. O categorizer
    # (mappings/categorias.yaml) aplica regras regex amplas como
    # `NU.PAGAMENT|BANCO.SANTA` que marcam `tipo: Transferência Interna`
    # em qualquer PIX que passe por Nubank/Santander como intermediário
    # financeiro -- inclusive PIX para terceiros (DEIVID, JOAO, etc.).
    # Esta passagem reverte a marcação quando a descrição NÃO casa
    # identidade do casal e não bate regra operacional legítima (pagamento
    # de fatura do próprio banco, CDB, agência 6450). Rodar DEPOIS do
    # categorizer é crítico: o categorizer pode reintroduzir falsos-
    # positivos se rodarmos antes.
    ctx["transacoes"] = _reclassificar_ti_orfas(ctx["transacoes"])
    # 6c. Promover variantes curtas para Transferência Interna (Sprint 82)
    # -- rede de captura simétrica. A Sprint 68b cobriu o rigoroso
    # (nome_aceitos completo + CPF); a 82 cobre formas abreviadas que só
    # são seguras sob contexto bancário (ex: "pessoa_b" no Itaú com
    # marcador PIX/TRANSF + data DD/MM, "ANDRE SILVA BATISTA FARIAS"
    # sem o "DA"). Roda DEPOIS do reclassificar para não reintroduzir
    # falsos-positivos que acabaram de ser degradados.
    ctx["transacoes"] = _promover_variantes_para_ti(ctx["transacoes"])
    n_ti = sum(1 for t in ctx["transacoes"] if t.get("tipo") == "Transferência Interna")
    return {"n_total": len(ctx["transacoes"]), "n_ti": n_ti}


def fase_aplicar_tags_irpf(ctx: dict[str, Any]) -> dict[str, Any]:
    """Fase 7: marca transações elegíveis a IRPF (saúde, educação, doação...)."""
    ctx["transacoes"] = aplicar_tags_irpf(ctx["transacoes"])
    n_tagged = sum(1 for t in ctx["transacoes"] if t.get("tag_irpf"))
    return {"n_total": len(ctx["transacoes"]), "n_tagged": n_tagged}


def fase_filtrar_por_mes(ctx: dict[str, Any]) -> dict[str, Any]:
    """Fase 8: filtra para mês específico se `--mes` foi passado sem `--tudo`."""
    mes = ctx.get("mes")
    processar_tudo = ctx.get("processar_tudo", False)
    if mes and not processar_tudo:
        ctx["transacoes_filtradas"] = _filtrar_por_mes(ctx["transacoes"], mes)
        logger.info("Filtrado para %s: %d transações", mes, len(ctx["transacoes_filtradas"]))
        return {"filtrado": True, "n_apos_filtro": len(ctx["transacoes_filtradas"])}
    ctx["transacoes_filtradas"] = ctx["transacoes"]
    return {"filtrado": False, "n_apos_filtro": len(ctx["transacoes_filtradas"])}


def fase_ordenar(ctx: dict[str, Any]) -> dict[str, Any]:
    """Fase 9: ordena por data e computa identificadores canônicos (hash de
    transação) para ativar `Doc?` no Extrato em runtime real (Sprint 87b)."""
    ctx["transacoes_filtradas"].sort(key=lambda t: t.get("data", ""))
    # 9b. Computar identificador canônico (mesmo hash dos nodes `transacao` do  # noqa: accent
    # grafo) para ativar `Doc?` no Extrato em runtime real (Sprint 87b).
    from src.graph.migracao_inicial import hash_transacao_do_tx

    n_ident = 0
    for tx in ctx["transacoes_filtradas"]:
        if tx.get("identificador"):
            continue  # contrato defensivo: não sobrescrever se já existir
        ident = hash_transacao_do_tx(tx)
        if ident is not None:
            tx["identificador"] = ident
            n_ident += 1
    return {"n_ordenadas": len(ctx["transacoes_filtradas"]), "n_identificadores": n_ident}


def fase_holerites(ctx: dict[str, Any]) -> dict[str, Any]:
    """Fase 10: ingere holerites (contracheques) como fonte extra de renda.
    P3.2 (auditoria 2026-04-23): também grava cada holerite como node
    `documento` no grafo (fecha ADR-20 tracking para folha)."""
    global _ESTAGIO_ATUAL
    _ESTAGIO_ATUAL = "holerites"
    from src.graph.db import GrafoDB, caminho_padrao

    caminho_grafo_hol = caminho_padrao()
    contracheques: list[dict] = []
    grafo_ok = False
    if caminho_grafo_hol.exists():
        # Sprint INFRA-PIPELINE-TRANSACIONALIDADE: criar_schema é DDL
        # idempotente (fora da transação); processar_holerites é a fase
        # mutadora que precisa de rollback granular se crashar.
        with GrafoDB(caminho_grafo_hol) as grafo_hol:
            grafo_hol.criar_schema()
            try:
                with grafo_hol.transaction():
                    contracheques = processar_holerites(
                        DIR_RAW / "andre" / "holerites", grafo=grafo_hol
                    )
                    grafo_ok = True
            except Exception as erro:
                _registrar_falha_pipeline_estruturada("processar_holerites", erro)
                logger.warning("Holerites com grafo falhou: %s", erro)
                contracheques = processar_holerites(DIR_RAW / "andre" / "holerites")
    else:
        contracheques = processar_holerites(DIR_RAW / "andre" / "holerites")
    ctx["contracheques"] = contracheques
    return {"n_contracheques": len(contracheques), "grafo_ok": grafo_ok}


def fase_gerar_xlsx(ctx: dict[str, Any]) -> dict[str, Any]:
    """Fase 11: gera planilha consolidada (8 abas canônicas)."""
    global _ESTAGIO_ATUAL
    _ESTAGIO_ATUAL = "gerar_xlsx"
    mes = ctx.get("mes")
    ano = mes[:4] if mes else str(datetime.now().year)
    caminho_xlsx = DIR_OUTPUT / f"ouroboros_{ano}.xlsx"
    gerar_xlsx(ctx["transacoes_filtradas"], caminho_xlsx, CONTROLE_ANTIGO, ctx["contracheques"])
    ctx["caminho_xlsx"] = caminho_xlsx
    return {"caminho_xlsx": str(caminho_xlsx), "ano": ano}


def fase_gerar_relatorios(ctx: dict[str, Any]) -> dict[str, Any]:
    """Fase 12: gera relatórios markdown auxiliares. Usa transações completas
    (para projeções corretas) mas filtra geração apenas para o mês solicitado."""
    mes = ctx.get("mes")
    processar_tudo = ctx.get("processar_tudo", False)
    meses_filtro = [mes] if (mes and not processar_tudo) else None
    gerar_relatorios(ctx["transacoes"], DIR_OUTPUT, meses_filtro=meses_filtro)
    return {"meses_filtro": meses_filtro, "n_transacoes": len(ctx["transacoes"])}


def fase_linking_documentos(ctx: dict[str, Any]) -> dict[str, Any]:
    """Fase 13: linking de documentos fiscais às transações bancárias
    (Sprint 48). Roda apenas se grafo SQLite já existir. Ausência de grafo
    não é erro -- pipeline principal do XLSX segue funcionando sem ele."""
    global _ESTAGIO_ATUAL
    _ESTAGIO_ATUAL = "linking_documentos"
    _executar_linking_documentos()
    return {"chamada": True}


def fase_er_produtos(ctx: dict[str, Any]) -> dict[str, Any]:
    """Fase 14: entity resolution de produtos (Sprint 49). Agrupa itens
    equivalentes em nodes `produto_canonico`. Roda depois do linking para
    que itens recém-linkados ja entrem no agregado. Ausência de grafo é no-op."""
    global _ESTAGIO_ATUAL
    _ESTAGIO_ATUAL = "er_produtos"
    _executar_er_produtos()
    return {"chamada": True}


def fase_item_categorizer(ctx: dict[str, Any]) -> dict[str, Any]:
    """Fase 15: categorização de itens (Sprint 50). Aplica regex de
    `mappings/categorias_item.yaml` aos nodes `item`, criando aresta
    `categoria_de` -> `categoria`. Itens em "Outros" com freq>=3 geram
    proposta MD para revisão supervisor."""
    global _ESTAGIO_ATUAL
    _ESTAGIO_ATUAL = "item_categorizer"
    _executar_item_categorizer()
    return {"chamada": True}


def fase_skill_d7_log(ctx: dict[str, Any]) -> dict[str, Any]:
    """Fase 16: snapshot do classificador D7 (Sprint INFRA-SKILLS-D7-LOG).
    Gera `data/output/skill_d7_log.json` consumido por dashboard. Falha-soft
    (sem grafo == no-op)."""
    global _ESTAGIO_ATUAL
    _ESTAGIO_ATUAL = "skill_d7_log"
    _executar_skill_d7_log()
    return {"chamada": True}


def fase_dossie_snapshot(ctx: dict[str, Any]) -> dict[str, Any]:
    """Fase 17: snapshot do dossie de graduacao por tipo documental
    (2026-05-13). Materializa o estado vivo de cada tipo
    (PENDENTE/CALIBRANDO/GRADUADO/REGREDINDO) consumido pelo dashboard
    `graduacao_tipos.py`. Ver `docs/CICLO_GRADUACAO_OPERACIONAL.md`."""
    global _ESTAGIO_ATUAL
    _ESTAGIO_ATUAL = "dossie_snapshot"
    _executar_dossie_snapshot()
    return {"chamada": True}


# Lista canônica de fases em ordem de execução. Pipeline serial depende
# desta ordem: cada fase pode mutar `ctx` e fases posteriores leem as
# chaves populadas pelas anteriores.
FASES: list[tuple[str, Any]] = [
    ("descobrir_extratores", fase_descobrir_extratores),
    ("escanear_arquivos", fase_escanear_arquivos),
    ("extrair", fase_extrair),
    ("importar_historico", fase_importar_historico),
    ("deduplicar", fase_deduplicar),
    ("categorizar", fase_categorizar),
    ("aplicar_tags_irpf", fase_aplicar_tags_irpf),
    ("filtrar_por_mes", fase_filtrar_por_mes),
    ("ordenar", fase_ordenar),
    ("holerites", fase_holerites),
    ("gerar_xlsx", fase_gerar_xlsx),
    ("gerar_relatorios", fase_gerar_relatorios),
    ("linking_documentos", fase_linking_documentos),
    ("er_produtos", fase_er_produtos),
    ("item_categorizer", fase_item_categorizer),
    ("skill_d7_log", fase_skill_d7_log),
    ("dossie_snapshot", fase_dossie_snapshot),
]


def _gravar_log_fases(stats_por_fase: list[dict[str, Any]]) -> Path | None:
    """Grava log estruturado JSON com stats de cada fase. Best-effort: erro
    de gravação NÃO aborta o pipeline (loga warning e segue). Retorna o path
    do arquivo gerado, ou None se a gravação falhou."""
    import json as _json

    try:
        DIR_OUTPUT.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        destino = DIR_OUTPUT / f"pipeline_fases_{ts}.json"
        payload = {
            "ts": ts,
            "n_fases": len(stats_por_fase),
            "fases": stats_por_fase,
        }
        destino.write_text(_json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        return destino
    except Exception as erro:
        logger.warning("Falha ao gravar log estruturado de fases: %s", erro)
        return None


def _executar_corpo_pipeline(mes: str | None, processar_tudo: bool) -> None:
    """Corpo do pipeline (orquestrador linear sobre `FASES`). Sprint
    INFRA-PIPELINE-FASES-ISOLADAS (2026-05-17): cada fase é função pura
    sobre `ctx` mutável, devolvendo dict de stats para log estruturado.
    `_ESTAGIO_ATUAL` continua sendo atualizado dentro de cada fase para
    diagnóstico do log de falha em caso de crash."""
    ctx: dict[str, Any] = {"mes": mes, "processar_tudo": processar_tudo}
    stats_por_fase: list[dict[str, Any]] = []
    for nome, fn in FASES:
        inicio = time.monotonic()
        stats = fn(ctx)
        stats["fase"] = nome
        stats["duracao_s"] = round(time.monotonic() - inicio, 4)
        stats_por_fase.append(stats)
    _gravar_log_fases(stats_por_fase)
    logger.info("=== Pipeline concluído ===")
    logger.info("XLSX: %s", ctx.get("caminho_xlsx"))
    logger.info("Relatórios: %s", DIR_OUTPUT)


def _executar_backfill_metadata() -> int:
    """Rota administrativa da Sprint 87.5: backfill de arquivo_original.

    Isola-se do pipeline regular; chama apenas o módulo dedicado e sai.
    Retorna exit code (0 sempre que conclui sem exceção)."""
    from src.graph.backfill_arquivo_original import backfill_arquivo_original
    from src.graph.db import GrafoDB, caminho_padrao

    db = GrafoDB(caminho_padrao())
    try:
        stats = backfill_arquivo_original(db)
    finally:
        db.fechar()
    logger.info("backfill-metadata stats: %s", stats)
    # Saída legível no stdout do operador; pipeline não produz outros prints.
    for chave, valor in stats.items():
        sys.stdout.write(f"{chave}: {valor}\n")
    return 0


def main(argv: list[str] | None = None) -> None:
    """Entrypoint CLI."""
    parser = argparse.ArgumentParser(description="Pipeline ETL Financeiro")
    parser.add_argument("--mes", type=str, help="Mês para processar (YYYY-MM)")
    parser.add_argument("--tudo", action="store_true", help="Processar todos os dados")
    parser.add_argument(
        "--backfill-metadata",
        action="store_true",
        help="Rota administrativa: preenche arquivo_original em nodes documento (Sprint 87.5)",
    )
    parser.add_argument(
        "--restore-grafo",
        type=str,
        metavar="TIMESTAMP",
        help=(
            "Restaura grafo de backup data/output/backup/grafo_<TIMESTAMP>.sqlite "
            "(formato YYYY-MM-DD_HHMMSS)."
        ),
    )
    args = parser.parse_args(argv)

    if args.restore_grafo:
        sys.exit(_restaurar_grafo_de_backup(args.restore_grafo))

    if args.backfill_metadata:
        sys.exit(_executar_backfill_metadata())

    if not args.mes and not args.tudo:
        parser.print_help()
        sys.exit(1)

    executar(mes=args.mes, processar_tudo=args.tudo)


if __name__ == "__main__":
    main()


# "A verdadeira sabedoria está em reconhecer a própria ignorância." -- Sócrates
