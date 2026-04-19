"""Geração do XLSX consolidado com 8 abas."""

from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.utils.logger import configurar_logger

logger = configurar_logger("xlsx_writer")

COLUNAS_EXTRATO = [
    "data",
    "valor",
    "forma_pagamento",
    "local",
    "quem",
    "categoria",
    "classificacao",
    "banco_origem",
    "tipo",
    "mes_ref",
    "tag_irpf",
    "obs",
]

AVISO_SNAPSHOT = (
    "[Snapshot histórico 2023 -- dados não são atualizados automaticamente. "
    "Reabilitação prevista para a Sprint 24 (automação bancária).]"
)

HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
VALOR_FORMAT = "#,##0.00"


def _aplicar_estilo_header(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    linha: int = 1,
) -> None:
    """Aplica estilo visual nos cabeçalhos da linha informada (default: 1)."""
    for cell in ws[linha]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _ajustar_largura_colunas(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Ajusta a largura das colunas pelo conteúdo."""
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        largura = min(max_len + 3, 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = largura


def _criar_aba_extrato(wb: openpyxl.Workbook, transacoes: list[dict]) -> None:
    """Cria a aba principal de extrato."""
    ws = wb.create_sheet("extrato")

    # Cabeçalhos
    for col_idx, col_nome in enumerate(COLUNAS_EXTRATO, 1):
        ws.cell(row=1, column=col_idx, value=col_nome)

    # Dados
    for row_idx, t in enumerate(transacoes, 2):
        for col_idx, col_nome in enumerate(COLUNAS_EXTRATO, 1):
            valor = t.get(col_nome)
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)

            if col_nome == "valor":
                cell.number_format = VALOR_FORMAT
            elif col_nome == "data" and isinstance(valor, (date, datetime)):
                cell.number_format = "DD/MM/YYYY"

    _aplicar_estilo_header(ws)
    _ajustar_largura_colunas(ws)

    # Filtro automático
    if len(transacoes) > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUNAS_EXTRATO))}{len(transacoes) + 1}"

    logger.info("Aba 'extrato': %d transações escritas", len(transacoes))


def _criar_aba_renda(
    wb: openpyxl.Workbook,
    transacoes: list[dict],
    contracheques: Optional[list[dict]] = None,
) -> None:
    """Cria a aba de renda mensal.

    Prioriza holerites extraídos (com bruto/INSS/IRRF/VR-VA reais) quando
    disponíveis. Para meses sem holerite, recai no comportamento antigo de
    inferir receita bruta das transações marcadas como "Receita" no extrato.
    """
    ws = wb.create_sheet("renda")
    colunas = ["mes_ref", "fonte", "bruto", "inss", "irrf", "vr_va", "liquido", "banco"]

    for col_idx, col_nome in enumerate(colunas, 1):
        ws.cell(row=1, column=col_idx, value=col_nome)

    contracheques = contracheques or []
    meses_com_holerite = {c["mes_ref"] for c in contracheques}

    row_idx = 2

    for c in sorted(contracheques, key=lambda x: (x["mes_ref"], x["fonte"])):
        ws.cell(row=row_idx, column=1, value=c["mes_ref"])
        ws.cell(row=row_idx, column=2, value=c["fonte"])
        ws.cell(row=row_idx, column=3, value=c.get("bruto", 0))
        ws.cell(row=row_idx, column=4, value=c.get("inss", 0))
        ws.cell(row=row_idx, column=5, value=c.get("irrf", 0))
        ws.cell(row=row_idx, column=6, value=c.get("vr_va", 0))
        ws.cell(row=row_idx, column=7, value=c.get("liquido", 0))
        ws.cell(row=row_idx, column=8, value=c.get("banco", ""))
        for col in (3, 4, 5, 6, 7):
            ws.cell(row=row_idx, column=col).number_format = VALOR_FORMAT
        row_idx += 1

    receitas_por_mes: dict[str, list[dict]] = {}
    for t in transacoes:
        if t.get("tipo") == "Receita":
            mes = t.get("mes_ref", "")
            if mes in meses_com_holerite:
                continue
            receitas_por_mes.setdefault(mes, []).append(t)

    for mes in sorted(receitas_por_mes.keys()):
        for t in receitas_por_mes[mes]:
            ws.cell(row=row_idx, column=1, value=mes)
            ws.cell(row=row_idx, column=2, value=t.get("local", ""))
            ws.cell(row=row_idx, column=3, value=t.get("valor", 0))
            ws.cell(row=row_idx, column=7, value=t.get("valor", 0))
            ws.cell(row=row_idx, column=8, value=t.get("banco_origem", ""))
            ws.cell(row=row_idx, column=3).number_format = VALOR_FORMAT
            ws.cell(row=row_idx, column=7).number_format = VALOR_FORMAT
            row_idx += 1

    _aplicar_estilo_header(ws)
    _ajustar_largura_colunas(ws)
    logger.info(
        "Aba 'renda': %d linhas (%d holerites + %d receitas inferidas)",
        row_idx - 2,
        len(contracheques),
        row_idx - 2 - len(contracheques),
    )


def _criar_aba_resumo_mensal(wb: openpyxl.Workbook, transacoes: list[dict]) -> None:
    """Cria a aba de resumo mensal gerada automaticamente."""
    ws = wb.create_sheet("resumo_mensal")
    colunas = [
        "mes_ref",
        "receita_total",
        "despesa_total",
        "saldo",
        "top_categoria",
        "top_gasto",
        "total_obrigatorio",
        "total_superfluo",
        "total_questionavel",
    ]

    for col_idx, col_nome in enumerate(colunas, 1):
        ws.cell(row=1, column=col_idx, value=col_nome)

    # Agrupar por mês
    dados_mes: dict[str, dict] = {}
    for t in transacoes:
        if t.get("tipo") == "Transferência Interna":
            continue

        mes = t.get("mes_ref", "")
        if mes not in dados_mes:
            dados_mes[mes] = {
                "receita": 0,
                "despesa": 0,
                "categorias": {},
                "classificacoes": {"Obrigatório": 0, "Supérfluo": 0, "Questionável": 0},
            }

        valor = t.get("valor", 0)
        if t.get("tipo") == "Receita":
            dados_mes[mes]["receita"] += valor
        elif t.get("tipo") in ("Despesa", "Imposto"):
            dados_mes[mes]["despesa"] += valor
            cat = t.get("categoria", "Outros")
            dados_mes[mes]["categorias"][cat] = dados_mes[mes]["categorias"].get(cat, 0) + valor
            clf = t.get("classificacao", "Questionável")
            if clf in dados_mes[mes]["classificacoes"]:
                dados_mes[mes]["classificacoes"][clf] += valor

    row_idx = 2
    for mes in sorted(dados_mes.keys()):
        d = dados_mes[mes]
        top_cat = max(d["categorias"], key=d["categorias"].get) if d["categorias"] else ""
        top_val = d["categorias"].get(top_cat, 0)

        ws.cell(row=row_idx, column=1, value=mes)
        ws.cell(row=row_idx, column=2, value=d["receita"])
        ws.cell(row=row_idx, column=3, value=d["despesa"])
        ws.cell(row=row_idx, column=4, value=d["receita"] - d["despesa"])
        ws.cell(row=row_idx, column=5, value=top_cat)
        ws.cell(row=row_idx, column=6, value=f"R$ {top_val:,.2f}")
        ws.cell(row=row_idx, column=7, value=d["classificacoes"]["Obrigatório"])
        ws.cell(row=row_idx, column=8, value=d["classificacoes"]["Supérfluo"])
        ws.cell(row=row_idx, column=9, value=d["classificacoes"]["Questionável"])

        for c in [2, 3, 4, 7, 8, 9]:
            ws.cell(row=row_idx, column=c).number_format = VALOR_FORMAT

        row_idx += 1

    _aplicar_estilo_header(ws)
    _ajustar_largura_colunas(ws)
    logger.info("Aba 'resumo_mensal': %d meses", row_idx - 2)


def _criar_aba_dividas(
    wb: openpyxl.Workbook,
    caminho_historico: Optional[Path] = None,
) -> None:
    """Cria a aba de dívidas ativas, importando do histórico se existir."""
    ws = wb.create_sheet("dividas_ativas")
    colunas = ["mes_ref", "custo", "valor", "status", "vencimento", "quem", "recorrente", "obs"]

    ws.cell(row=1, column=1, value=AVISO_SNAPSHOT)
    for col_idx, col_nome in enumerate(colunas, 1):
        ws.cell(row=2, column=col_idx, value=col_nome)

    row_idx = 3
    if caminho_historico and caminho_historico.exists():
        try:
            wb_hist = openpyxl.load_workbook(caminho_historico)
            if "Dívidas Ativas" in wb_hist.sheetnames:
                ws_hist = wb_hist["Dívidas Ativas"]
                for i, row in enumerate(ws_hist.iter_rows(min_row=2, values_only=True)):
                    valores = [v for v in row if v is not None]
                    if not valores:
                        continue
                    custo = row[0] if len(row) > 0 else ""
                    valor = row[1] if len(row) > 1 else 0
                    status = row[2] if len(row) > 2 else ""
                    obs = row[3] if len(row) > 3 else ""

                    ws.cell(row=row_idx, column=1, value=date.today().strftime("%Y-%m"))
                    ws.cell(row=row_idx, column=2, value=custo)
                    ws.cell(
                        row=row_idx, column=3, value=valor if isinstance(valor, (int, float)) else 0
                    )
                    ws.cell(row=row_idx, column=4, value=status)
                    ws.cell(row=row_idx, column=8, value=obs)

                    if isinstance(valor, (int, float)):
                        ws.cell(row=row_idx, column=3).number_format = VALOR_FORMAT

                    row_idx += 1

                logger.info("Aba 'dividas_ativas': importadas %d linhas do histórico", row_idx - 3)
        except Exception as e:
            logger.warning("Erro ao importar dívidas do histórico: %s", e)

    _aplicar_estilo_header(ws, linha=2)
    _ajustar_largura_colunas(ws)


def _criar_aba_inventario(
    wb: openpyxl.Workbook,
    caminho_historico: Optional[Path] = None,
) -> None:
    """Cria a aba de inventário, importando e melhorando do histórico."""
    ws = wb.create_sheet("inventario")
    colunas = [
        "bem",
        "valor_aquisicao",
        "vida_util_anos",
        "depreciacao_anual",
        "perda_mensal",
    ]

    ws.cell(row=1, column=1, value=AVISO_SNAPSHOT)
    for col_idx, col_nome in enumerate(colunas, 1):
        ws.cell(row=2, column=col_idx, value=col_nome)

    row_idx = 3
    if caminho_historico and caminho_historico.exists():
        try:
            wb_hist = openpyxl.load_workbook(caminho_historico)
            if "Inventário" in wb_hist.sheetnames:
                ws_hist = wb_hist["Inventário"]
                for row in ws_hist.iter_rows(min_row=2, values_only=True):
                    bem = row[0] if len(row) > 0 else None
                    if bem is None or not isinstance(bem, str):
                        continue
                    valor_aq = row[1] if len(row) > 1 and isinstance(row[1], (int, float)) else 0
                    vida_util = row[2] if len(row) > 2 and isinstance(row[2], (int, float)) else 2
                    dep_anual = row[3] if len(row) > 3 and isinstance(row[3], (int, float)) else 0.1
                    perda = (valor_aq * dep_anual) / vida_util if vida_util > 0 else 0

                    ws.cell(row=row_idx, column=1, value=bem)
                    ws.cell(row=row_idx, column=2, value=valor_aq)
                    ws.cell(row=row_idx, column=3, value=vida_util)
                    ws.cell(row=row_idx, column=4, value=dep_anual)
                    ws.cell(row=row_idx, column=5, value=round(perda, 2))

                    ws.cell(row=row_idx, column=2).number_format = VALOR_FORMAT
                    ws.cell(row=row_idx, column=5).number_format = VALOR_FORMAT
                    row_idx += 1

                logger.info("Aba 'inventario': importados %d bens do histórico", row_idx - 3)
        except Exception as e:
            logger.warning("Erro ao importar inventário do histórico: %s", e)

    _aplicar_estilo_header(ws, linha=2)
    _ajustar_largura_colunas(ws)


def _criar_aba_prazos(
    wb: openpyxl.Workbook,
    caminho_historico: Optional[Path] = None,
) -> None:
    """Cria a aba de prazos de vencimento."""
    ws = wb.create_sheet("prazos")
    colunas = ["conta", "dia_vencimento", "banco_pagamento", "auto_debito"]

    ws.cell(row=1, column=1, value=AVISO_SNAPSHOT)
    for col_idx, col_nome in enumerate(colunas, 1):
        ws.cell(row=2, column=col_idx, value=col_nome)

    row_idx = 3
    if caminho_historico and caminho_historico.exists():
        try:
            wb_hist = openpyxl.load_workbook(caminho_historico)
            if "Prazos" in wb_hist.sheetnames:
                ws_hist = wb_hist["Prazos"]
                for row in ws_hist.iter_rows(min_row=8, values_only=True):
                    # Dados podem estar nas colunas C e D (índices 2 e 3)
                    conta = (
                        row[2] if len(row) > 2 and row[2] else (row[0] if len(row) > 0 else None)
                    )
                    dia = row[3] if len(row) > 3 and row[3] else (row[1] if len(row) > 1 else None)
                    if conta and dia and str(conta).strip() != "Dias":
                        ws.cell(row=row_idx, column=1, value=conta)
                        ws.cell(row=row_idx, column=2, value=dia)
                        row_idx += 1

                logger.info("Aba 'prazos': importados %d prazos do histórico", row_idx - 3)
        except Exception as e:
            logger.warning("Erro ao importar prazos do histórico: %s", e)

    _aplicar_estilo_header(ws, linha=2)
    _ajustar_largura_colunas(ws)


def _criar_aba_irpf(wb: openpyxl.Workbook, transacoes: list[dict]) -> None:
    """Cria a aba de dados relevantes para IRPF."""
    ws = wb.create_sheet("irpf")
    colunas = ["ano", "tipo", "fonte", "cnpj_cpf", "valor", "mes"]

    for col_idx, col_nome in enumerate(colunas, 1):
        ws.cell(row=1, column=col_idx, value=col_nome)

    row_idx = 2
    for t in transacoes:
        tag = t.get("tag_irpf")
        if tag is None:
            continue

        data_t = t.get("data")
        ano = data_t.year if hasattr(data_t, "year") else 0

        ws.cell(row=row_idx, column=1, value=ano)
        ws.cell(row=row_idx, column=2, value=tag)
        ws.cell(row=row_idx, column=3, value=t.get("local", ""))
        ws.cell(row=row_idx, column=4, value=t.get("cnpj_cpf", ""))
        ws.cell(row=row_idx, column=5, value=t.get("valor", 0))
        ws.cell(row=row_idx, column=6, value=t.get("mes_ref", ""))
        ws.cell(row=row_idx, column=5).number_format = VALOR_FORMAT
        row_idx += 1

    _aplicar_estilo_header(ws)
    _ajustar_largura_colunas(ws)
    logger.info("Aba 'irpf': %d registros com tag IRPF", row_idx - 2)


def _criar_aba_analise(wb: openpyxl.Workbook, transacoes: list[dict]) -> None:
    """Cria a aba de análise com insights quantitativos e qualitativos.

    A aba é marcada como DEPRECATED: produz apenas totais e contagens, sem
    análise interpretativa. Mantida para compatibilidade até a implementação
    do resumo narrativo baseado em LLM (Sprint 33).
    """
    from collections import Counter

    ws = wb.create_sheet("analise")
    ws.column_dimensions["A"].width = 80
    negrito = Font(bold=True, size=12)
    subtitulo = Font(bold=True, size=11, color="2F5496")
    alerta = Font(bold=True, size=11, color="B00020")

    despesas = [t for t in transacoes if t.get("tipo") in ("Despesa", "Imposto")]
    receitas = [t for t in transacoes if t.get("tipo") == "Receita"]
    total_desp = sum(t["valor"] for t in despesas)
    total_rec = sum(t["valor"] for t in receitas)
    saldo = total_rec - total_desp

    meses = sorted({t.get("mes_ref", "") for t in transacoes if t.get("mes_ref")})
    n_meses = len(meses) or 1

    linhas: list[tuple[str, Optional[Font]]] = [
        (
            "[DEPRECATED -- apenas totais, sem análise interpretativa. "
            "Resumo narrativo diagnóstico em implementação na Sprint 33.]",
            alerta,
        ),
        ("", None),
        ("PANORAMA GERAL", negrito),
        (
            f"Período: {meses[0] if meses else '?'} a "
            f"{meses[-1] if meses else '?'} ({n_meses} meses)",
            None,
        ),
        (
            f"Transações: {len(transacoes)} ({len(receitas)} receitas, {len(despesas)} despesas)",
            None,
        ),
        (f"Receita total: R$ {total_rec:,.2f} (média R$ {total_rec / n_meses:,.2f}/mês)", None),
        (f"Despesa total: R$ {total_desp:,.2f} (média R$ {total_desp / n_meses:,.2f}/mês)", None),
        (f"Saldo acumulado: R$ {saldo:,.2f}", None),
        (f"Taxa de poupança: {(saldo / total_rec * 100) if total_rec > 0 else 0:.1f}%", None),
        ("", None),
    ]

    # Top 10 categorias por gasto
    cat_totais = Counter()
    for t in despesas:
        cat_totais[t.get("categoria", "?")] += t["valor"]
    top10 = cat_totais.most_common(10)

    linhas.append(("TOP 10 CATEGORIAS (por valor total)", subtitulo))
    for i, (cat, val) in enumerate(top10, 1):
        pct = val / total_desp * 100 if total_desp > 0 else 0
        linhas.append((f"  {i}. {cat}: R$ {val:,.2f} ({pct:.1f}%)", None))
    linhas.append(("", None))

    # Classificação
    clf_totais = Counter()
    for t in despesas:
        clf_totais[t.get("classificacao", "?")] += t["valor"]

    linhas.append(("CLASSIFICAÇÃO DE DESPESAS", subtitulo))
    for clf in ["Obrigatório", "Questionável", "Supérfluo"]:
        val = clf_totais.get(clf, 0)
        pct = val / total_desp * 100 if total_desp > 0 else 0
        linhas.append((f"  {clf}: R$ {val:,.2f} ({pct:.1f}%)", None))
    linhas.append(("", None))

    # Análise por pessoa
    pessoa_desp: dict[str, float] = {}
    pessoa_rec: dict[str, float] = {}
    for t in despesas:
        p = t.get("quem", "?")
        pessoa_desp[p] = pessoa_desp.get(p, 0) + t["valor"]
    for t in receitas:
        p = t.get("quem", "?")
        pessoa_rec[p] = pessoa_rec.get(p, 0) + t["valor"]

    linhas.append(("BALANÇO POR PESSOA", subtitulo))
    for p in sorted(set(list(pessoa_desp.keys()) + list(pessoa_rec.keys()))):
        rec = pessoa_rec.get(p, 0)
        desp = pessoa_desp.get(p, 0)
        saldo_p = rec - desp
        txt_p = f"  {p}: R$ {rec:,.2f} rec | R$ {desp:,.2f} desp | R$ {saldo_p:,.2f}"
        linhas.append((txt_p, None))
    linhas.append(("", None))

    # Evolução mensal (últimos 6 meses)
    ultimos_6 = meses[-6:] if len(meses) >= 6 else meses
    linhas.append(("EVOLUÇÃO MENSAL (últimos meses)", subtitulo))
    for m in ultimos_6:
        rec_m = sum(t["valor"] for t in receitas if t.get("mes_ref") == m)
        desp_m = sum(t["valor"] for t in despesas if t.get("mes_ref") == m)
        saldo_m = rec_m - desp_m
        sinal = "+" if saldo_m >= 0 else ""
        txt_m = f"  {m}: R$ {rec_m:,.2f} rec | R$ {desp_m:,.2f} desp | {sinal}R$ {saldo_m:,.2f}"
        linhas.append((txt_m, None))
    linhas.append(("", None))

    # Anomalias (categorias com variação > 50% entre penúltimo e último mês)
    if len(meses) >= 2:
        penultimo, ultimo = meses[-2], meses[-1]
        cat_pen: Counter = Counter()
        cat_ult: Counter = Counter()
        for t in despesas:
            if t.get("mes_ref") == penultimo:
                cat_pen[t.get("categoria", "?")] += t["valor"]
            elif t.get("mes_ref") == ultimo:
                cat_ult[t.get("categoria", "?")] += t["valor"]

        anomalias: list[str] = []
        todas_cats = set(list(cat_pen.keys()) + list(cat_ult.keys()))
        for cat in sorted(todas_cats):
            v_pen = cat_pen.get(cat, 0)
            v_ult = cat_ult.get(cat, 0)
            if v_pen > 100 and v_ult > 100:
                variacao = (v_ult - v_pen) / v_pen * 100
                if abs(variacao) > 50:
                    sinal = "+" if variacao > 0 else ""
                    anomalias.append(
                        f"  {cat}: R$ {v_pen:,.2f} -> R$ {v_ult:,.2f} ({sinal}{variacao:.0f}%)"
                    )
            elif v_pen == 0 and v_ult > 200:
                anomalias.append(f"  {cat}: NOVO gasto de R$ {v_ult:,.2f} em {ultimo}")
            elif v_pen > 200 and v_ult == 0:
                anomalias.append(f"  {cat}: DESAPARECEU (era R$ {v_pen:,.2f} em {penultimo})")

        if anomalias:
            linhas.append((f"ANOMALIAS ({penultimo} vs {ultimo})", subtitulo))
            for a in anomalias:
                linhas.append((a, None))
            linhas.append(("", None))

    # Saúde financeira (categorias de saúde agrupadas)
    saude_cats = {"Farmácia", "Saúde", "Natação"}
    saude_total = sum(t["valor"] for t in despesas if t.get("categoria") in saude_cats)
    if saude_total > 0:
        pct_saude = saude_total / total_desp * 100 if total_desp > 0 else 0
        linhas.append(("CUSTOS DE SAÚDE (agrupado)", subtitulo))
        for cat in sorted(saude_cats):
            val = cat_totais.get(cat, 0)
            if val > 0:
                linhas.append((f"  {cat}: R$ {val:,.2f}", None))
        linhas.append(
            (f"  TOTAL SAÚDE: R$ {saude_total:,.2f} ({pct_saude:.1f}% das despesas)", None)
        )
        linhas.append(("", None))

    # Escrever na aba
    for i, (texto, fonte) in enumerate(linhas, 1):
        cell = ws.cell(row=i, column=1, value=texto)
        if fonte:
            cell.font = fonte


def gerar_xlsx(
    transacoes: list[dict],
    caminho_saida: Path,
    caminho_historico: Optional[Path] = None,
    contracheques: Optional[list[dict]] = None,
) -> Path:
    """Gera o XLSX completo com todas as 8 abas."""
    wb = openpyxl.Workbook()
    # Remover aba padrão
    wb.remove(wb.active)

    _criar_aba_extrato(wb, transacoes)
    _criar_aba_renda(wb, transacoes, contracheques)
    _criar_aba_dividas(wb, caminho_historico)
    _criar_aba_inventario(wb, caminho_historico)
    _criar_aba_prazos(wb, caminho_historico)
    _criar_aba_resumo_mensal(wb, transacoes)
    _criar_aba_irpf(wb, transacoes)
    _criar_aba_analise(wb, transacoes)

    caminho_saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho_saida)

    logger.info("XLSX gerado: %s (%d transações, 8 abas)", caminho_saida, len(transacoes))
    return caminho_saida


# "A riqueza não consiste em ter grandes posses, mas em ter poucas necessidades." -- Epicteto
