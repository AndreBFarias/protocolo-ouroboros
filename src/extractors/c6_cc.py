"""Extrator de extrato de conta corrente C6 Bank (formato XLSX criptografado)."""

import io
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

import msoffcrypto
import openpyxl

from src.extractors.base import ExtratorBase, Transacao
from src.transform.canonicalizer_casal import e_transferencia_do_casal
from src.utils.logger import configurar_logger
from src.utils.senhas import carregar_senhas_pdf

LINHA_CABECALHO: int = 8
COLUNAS: dict[str, int] = {
    "data_lancamento": 0,
    "data_contabil": 1,
    "titulo": 2,
    "descricao": 3,
    "entrada": 4,
    "saida": 5,
    "saldo": 6,
}

REGEX_FATURA_CARTAO: re.Pattern[str] = re.compile(
    r"PGTO\s+FAT\s+CARTAO|Fatura\s+de\s+cart[aã]o",
    re.IGNORECASE,
)

REGEX_CDB: re.Pattern[str] = re.compile(
    r"CDB\s+C6|LIM\.\s*GARANT",
    re.IGNORECASE,
)

REGEX_SALARIO: re.Pattern[str] = re.compile(
    r"RECEBIMENTO\s+SALARIO|PAGTO\s+SALARIO",
    re.IGNORECASE,
)

REGEX_IMPOSTO: re.Pattern[str] = re.compile(
    r"SEFAZ|DAS.*Simples|DARF|RECEITA\s*FED",
    re.IGNORECASE,
)


class ExtratorC6CC(ExtratorBase):
    """Extrai transações do XLSX de conta corrente C6 Bank.

    Arquivo criptografado com senha.
    Colunas: Data Lançamento, Data Contábil, Título, Descrição,
             Entrada(R$), Saída(R$), Saldo do Dia(R$)
    """

    def __init__(self, caminho: Path) -> None:
        super().__init__(caminho)
        self.logger = configurar_logger("ExtratorC6CC")
        self.senhas: list[str] = self._carregar_senhas()

    def pode_processar(self, caminho: Path) -> bool:
        """Verifica se é um XLSX de conta corrente C6."""
        if not caminho.suffix.lower() == ".xlsx":
            return False
        return "c6_cc" in str(caminho).lower()

    def extrair(self) -> list[Transacao]:
        """Extrai todas as transações do XLSX C6."""
        transacoes: list[Transacao] = []
        arquivos: list[Path] = self._listar_arquivos()

        if not arquivos:
            self.logger.warning("Nenhum arquivo XLSX encontrado em %s", self.caminho)
            return transacoes

        for arquivo in arquivos:
            self.logger.info("Processando extrato C6 CC: %s", arquivo.name)
            transacoes.extend(self._processar_arquivo(arquivo))

        self.logger.info("Total de transações extraídas (C6 CC): %d", len(transacoes))
        return transacoes

    def _listar_arquivos(self) -> list[Path]:
        """Lista todos os XLSX válidos no diretório."""
        if self.caminho.is_file():
            return [self.caminho] if self.pode_processar(self.caminho) else []
        return sorted(f for f in self.caminho.glob("*.xlsx") if self.pode_processar(f))

    def _processar_arquivo(self, arquivo: Path) -> list[Transacao]:
        """Processa um único XLSX e retorna lista de transações."""
        transacoes: list[Transacao] = []
        wb: Optional[openpyxl.Workbook] = self._abrir_workbook(arquivo)

        if wb is None:
            return transacoes

        ws = wb.active
        if ws is None:
            self.logger.error("Planilha ativa não encontrada em %s", arquivo.name)
            return transacoes

        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx <= LINHA_CABECALHO:
                continue

            transacao: Optional[Transacao] = self._parse_linha(list(row), arquivo)
            if transacao:
                transacoes.append(transacao)

        return transacoes

    def _abrir_workbook(self, arquivo: Path) -> Optional[openpyxl.Workbook]:
        """Abre o workbook, tentando descriptografar se necessário."""
        try:
            with open(arquivo, "rb") as f:
                ms = msoffcrypto.OfficeFile(f)
                if ms.is_encrypted():
                    return self._descriptografar(ms, arquivo)
                f.seek(0)
                return openpyxl.load_workbook(f)
        except Exception as erro:
            self.logger.error("Erro ao abrir %s: %s", arquivo.name, erro)
            return None

    def _descriptografar(
        self,
        ms: msoffcrypto.OfficeFile,
        arquivo: Path,
    ) -> Optional[openpyxl.Workbook]:
        """Tenta descriptografar com as senhas disponíveis."""
        for senha in self.senhas:
            try:
                ms.load_key(password=senha)
                buf: io.BytesIO = io.BytesIO()
                ms.decrypt(buf)
                buf.seek(0)
                self.logger.debug("Arquivo %s descriptografado com sucesso", arquivo.name)
                return openpyxl.load_workbook(buf)
            except Exception:
                continue

        self.logger.error("Nenhuma senha válida para %s", arquivo.name)
        return None

    def _parse_linha(
        self,
        valores: list[Any],
        arquivo: Path,
    ) -> Optional[Transacao]:
        """Converte uma linha da planilha em Transação."""
        try:
            if len(valores) < 7:
                return None

            data_str: Any = valores[COLUNAS["data_lancamento"]]
            titulo: Any = valores[COLUNAS["titulo"]]
            descricao_col: Any = valores[COLUNAS["descricao"]]
            entrada: Any = valores[COLUNAS["entrada"]]
            saida: Any = valores[COLUNAS["saida"]]

            if data_str is None or titulo is None:
                return None

            data_transacao: date = self._parse_data(data_str)
            if data_transacao is None:
                return None

            titulo_str: str = str(titulo).strip()
            descricao_str: str = str(descricao_col).strip() if descricao_col else ""
            descricao_completa: str = (
                f"{titulo_str} - {descricao_str}" if descricao_str else titulo_str
            )

            valor_entrada: float = float(entrada) if entrada else 0.0
            valor_saida: float = float(saida) if saida else 0.0

            if valor_entrada > 0:
                valor: float = valor_entrada
            elif valor_saida > 0:
                valor = -valor_saida
            else:
                return None

            forma_pagamento: str = self._inferir_forma_pagamento(titulo_str, descricao_str)
            tipo: str = self._classificar_tipo(titulo_str, descricao_str, valor)

            return Transacao(
                data=data_transacao,
                valor=valor,
                descricao=descricao_completa,
                banco_origem="C6",
                pessoa="pessoa_a",
                forma_pagamento=forma_pagamento,
                tipo=tipo,
                identificador=None,
                arquivo_origem=str(arquivo.name),
            )
        except (ValueError, TypeError, IndexError) as erro:
            self.logger.warning("Linha ignorada em %s: %s - %s", arquivo.name, valores, erro)
            return None

    @staticmethod
    def _parse_data(valor: Any) -> Optional[date]:
        """Converte valor de data (string DD/MM/YYYY ou datetime) para date."""
        if isinstance(valor, datetime):
            return valor.date()
        if isinstance(valor, date):
            return valor
        if isinstance(valor, str):
            try:
                return datetime.strptime(valor.strip(), "%d/%m/%Y").date()
            except ValueError:
                return None
        return None

    @staticmethod
    def _inferir_forma_pagamento(titulo: str, descricao: str) -> str:
        """Infere a forma de pagamento a partir do título e descrição."""
        texto: str = f"{titulo} {descricao}".upper()
        if "PIX" in texto:
            return "Pix"
        if "DEBITO DE CARTAO" in texto:
            return "Débito"
        if "BOLETO" in texto:
            return "Boleto"
        if "TED" in texto:
            return "Transferência"
        return "Transferência"

    @staticmethod
    def _classificar_tipo(titulo: str, descricao: str, valor: float) -> str:
        """Classifica o tipo da transação.

        Sprint 68b: substituiu `REGEX_VITORIA` e `REGEX_ANDRE_PROPRIO`
        (fragmentos genéricos como `ANDRE.*SILVA.*BATISTA` que casavam
        falsos-positivos com terceiros homônimos) pelo matcher formal
        `canonicalizer_casal.e_transferencia_do_casal`, que consulta
        `mappings/contas_casal.yaml`. `REGEX_FATURA_CARTAO` e `REGEX_CDB`
        preservados como regras operacionais legítimas (pagamento de
        fatura do próprio cartão; resgate/aplicação de CDB C6).
        """
        texto_completo: str = f"{titulo} {descricao}"

        if REGEX_SALARIO.search(texto_completo):
            return "Receita"

        if REGEX_FATURA_CARTAO.search(texto_completo):
            return "Transferência Interna"

        if REGEX_CDB.search(texto_completo):
            return "Transferência Interna"

        if e_transferencia_do_casal(texto_completo):
            return "Transferência Interna"

        if REGEX_IMPOSTO.search(texto_completo):
            return "Imposto"

        if valor > 0:
            return "Receita"

        return "Despesa"

    @staticmethod
    def _carregar_senhas() -> list[str]:
        """Carrega senhas via módulo centralizado."""
        return carregar_senhas_pdf()


# "O dinheiro é um bom servo, mas um mau mestre." -- Francis Bacon
