"""Sprint 87b: round-trip do identificador canônico do XLSX <-> grafo.

Cobre:
  - Presença de `identificador` em `COLUNAS_EXTRATO`.
  - Determinismo de `hash_transacao_canonico`.
  - Contrato defensivo de `hash_transacao_do_tx` (None quando falta local/banco).
  - Round-trip: XLSX gerado pelo writer contém a coluna com o hash esperado.
  - Fechamento existencial: ciclo Sprint 87.2 funciona em runtime real com
    grafo sintético contendo aresta `documento_de`.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Any

import openpyxl

from src.graph.migracao_inicial import hash_transacao_canonico, hash_transacao_do_tx
from src.load.xlsx_writer import COLUNAS_EXTRATO, gerar_xlsx


def _tx(data: date, valor: float, local: str, banco: str, **extra: Any) -> dict[str, Any]:
    base: dict[str, Any] = {
        "data": data,
        "valor": valor,
        "forma_pagamento": "Pix",
        "local": local,
        "quem": "Casal",
        "categoria": "Alimentação",
        "classificacao": "Obrigatório",
        "banco_origem": banco,
        "tipo": "Despesa",
        "mes_ref": data.strftime("%Y-%m"),
        "tag_irpf": None,
        "obs": None,
    }
    base.update(extra)
    return base


def test_colunas_extrato_inclui_identificador():
    assert "identificador" in COLUNAS_EXTRATO
    assert COLUNAS_EXTRATO[-1] == "identificador"  # spec: fim da lista


def test_hash_canonico_deterministico():
    h1 = hash_transacao_canonico(date(2026, 3, 15), 103.93, "SESC DF", "c6_cc")
    h2 = hash_transacao_canonico(date(2026, 3, 15), 103.93, "SESC DF", "c6_cc")
    assert h1 == h2
    assert len(h1) == 16


def test_hash_do_tx_retorna_none_quando_falta_local_ou_banco():
    assert hash_transacao_do_tx({"data": date(2026, 3, 15), "valor": 10.0, "local": "X"}) is None
    assert (
        hash_transacao_do_tx({"data": date(2026, 3, 15), "valor": 10.0, "banco_origem": "c6_cc"})
        is None
    )
    assert hash_transacao_do_tx({"valor": 10.0, "local": "X", "banco_origem": "c6_cc"}) is None


def test_round_trip_xlsx_contem_identificador(tmp_path: Path):
    txs = [
        _tx(date(2026, 3, 15), 103.93, "SESC DF", "c6_cc"),
        _tx(date(2026, 3, 16), 50.0, "Padaria X", "nubank_cc"),
        _tx(date(2026, 3, 17), 1234.56, "Loja Y", "itau"),
    ]
    # Pipeline injeta:
    for tx in txs:
        tx["identificador"] = hash_transacao_do_tx(tx)

    xlsx_path = tmp_path / "teste.xlsx"
    gerar_xlsx(txs, xlsx_path, Path("/inexistente.xlsx"), None)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["extrato"]
    cabecalho = [c.value for c in ws[1]]
    assert "identificador" in cabecalho
    idx_ident = cabecalho.index("identificador") + 1  # openpyxl usa 1-indexed
    for row_idx, tx in enumerate(txs, 2):
        valor_gravado = ws.cell(row=row_idx, column=idx_ident).value
        assert valor_gravado == tx["identificador"]
        # Simetria bit-a-bit com recomputação direta:
        esperado = hash_transacao_canonico(tx["data"], tx["valor"], tx["local"], tx["banco_origem"])
        assert valor_gravado == esperado
    wb.close()


def test_ciclo_87_2_funciona_em_runtime_com_grafo(tmp_path: Path):
    """Fechamento existencial da Sprint 87b.

    Quando o grafo tem aresta `documento_de` apontando para node `transacao`,  # noqa: accent
    e a coluna `identificador` no XLSX bate com `nome_canonico` do node,
    `transacoes_com_documento(db)` retorna o hash e `_marcar_tracking` marca OK.
    """
    from src.graph.db import GrafoDB
    from src.graph.queries import transacoes_com_documento

    db = GrafoDB(tmp_path / "grafo.sqlite")
    db.criar_schema()
    try:
        # Transação sintética -> hash canônico via helper público
        tx = _tx(date(2026, 3, 15), 103.93, "SESC DF", "c6_cc")
        hash_t = hash_transacao_do_tx(tx)
        assert hash_t is not None

        tx_id = db.upsert_node("transacao", hash_t, metadata={"data": str(tx["data"])})
        # Documento vinculado
        doc_id = db.upsert_node(
            "documento", "BOLETO-1", metadata={"tipo_documento": "boleto_servico"}
        )
        db.adicionar_edge(doc_id, tx_id, "documento_de")

        ids = transacoes_com_documento(db)
        assert hash_t in ids

        # Tx sem documento (outro hash) NÃO aparece
        outro = hash_transacao_canonico(date(2026, 3, 16), 50.0, "Padaria X", "nubank_cc")
        assert outro not in ids
    finally:
        db.fechar()


# "A verdade é filha do tempo, não da autoridade." -- Francis Bacon
