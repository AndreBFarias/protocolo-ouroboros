"""CLI: gera sugestões TF-IDF para transações "Outros" do XLSX.

Sprint CATEGORIZER-SUGESTAO-TFIDF (2026-05-16). Lê o XLSX
consolidado, aplica `categorizer_suggest.gerar_sugestoes`, grava
`data/output/sugestoes_categoria.json` para consumo do dashboard.

Uso:

    python scripts/sugerir_categorias.py
    python scripts/sugerir_categorias.py --top-k 3
    python scripts/sugerir_categorias.py --xlsx data/output/ouroboros_2026.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

from src.transform.categorizer_suggest import Transacao, gerar_sugestoes

_RAIZ = Path(__file__).resolve().parents[1]
PATH_XLSX_DEFAULT = _RAIZ / "data" / "output" / "ouroboros_2026.xlsx"
PATH_SUGESTOES = _RAIZ / "data" / "output" / "sugestoes_categoria.json"


def _carregar_transacoes_do_xlsx(xlsx: Path) -> list[Transacao]:
    """Lê aba `extrato` do XLSX. Schema canônico do projeto.

    Schema das colunas (XLSX `extrato`):
        0: data, 1: valor, 2: forma_pagamento, 3: local, 4: quem,
        5: categoria, 6: classificacao, 7: banco_origem, 8: tipo, ...

    Descrição da transação = `local` (col 3) — campo livre com nome do
    estabelecimento / contraparte do PIX / descrição do extrato.
    Categoria = col 5. ID sintético: posição + valor.
    Se XLSX ausente, retorna [].
    """
    if not xlsx.exists():
        return []
    try:
        import openpyxl  # type: ignore[import-untyped]
    except ImportError:
        return []
    wb = openpyxl.load_workbook(str(xlsx), read_only=True, data_only=True)
    if "extrato" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["extrato"]
    transacoes: list[Transacao] = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or len(row) < 6:
            continue
        descricao = str(row[3] or "")  # col[3] = local (descrição livre)
        categoria = str(row[5] or "")
        if not descricao or not categoria:
            continue
        transacoes.append(Transacao(id=f"row_{i + 2}", descricao=descricao, categoria=categoria))
    wb.close()
    return transacoes


def gerar_e_gravar(
    xlsx: Path | None = None,
    saida: Path | None = None,
    top_k: int = 5,
) -> dict:
    """Pipeline completo: lê XLSX, gera sugestões, grava JSON. Retorna payload."""
    xlsx_path = xlsx if xlsx is not None else PATH_XLSX_DEFAULT
    saida_path = saida if saida is not None else PATH_SUGESTOES
    transacoes = _carregar_transacoes_do_xlsx(xlsx_path)
    sugestoes = gerar_sugestoes(transacoes, top_k=top_k)
    payload = {
        "gerado_em": datetime.now(timezone.utc).isoformat(),
        "xlsx_origem": str(xlsx_path),
        "total_transacoes": len(transacoes),
        "total_outros": sum(1 for t in transacoes if t.categoria == "Outros"),
        "total_sugeridas": len(sugestoes),
        "sugestoes": sugestoes,
    }
    saida_path.parent.mkdir(parents=True, exist_ok=True)
    saida_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return payload


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Sugestor TF-IDF para Outros")
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=PATH_XLSX_DEFAULT,
        help="Caminho do XLSX (default: data/output/ouroboros_2026.xlsx)",
    )
    parser.add_argument(
        "--saida",
        type=Path,
        default=PATH_SUGESTOES,
        help="Caminho do JSON (default: data/output/sugestoes_categoria.json)",
    )
    parser.add_argument(
        "--top-k",
        type=int,
        default=5,
        help="Vizinhos top-K para votação (default: 5)",
    )
    args = parser.parse_args(argv)

    payload = gerar_e_gravar(xlsx=args.xlsx, saida=args.saida, top_k=args.top_k)
    sys.stdout.write(
        f"Sugestões em {args.saida}\n"
        f"  total transações: {payload['total_transacoes']}\n"
        f"  total Outros:     {payload['total_outros']}\n"
        f"  total sugeridas:  {payload['total_sugeridas']}\n"
    )
    alta = sum(1 for v in payload["sugestoes"].values() if v.get("confianca_top1", 0) >= 0.85)
    sys.stdout.write(f"  confiança >= 0.85: {alta} (auto-promoção candidata)\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())


# "Métrica de Outros é métrica de débito; sugestor é juro pago." -- principio
