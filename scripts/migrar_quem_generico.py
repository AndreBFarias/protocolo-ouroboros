"""Migra coluna ``quem`` dos XLSX em ``data/output/`` para identificadores genéricos.

Sprint MOB-bridge-1: schema canônico do XLSX agora usa
``pessoa_a``/``pessoa_b``/``casal`` na coluna ``quem`` (ADR-23).
XLSX antigos têm ``"André"``/``"Vitória"``/``"Casal"`` em registros
históricos. Este script faz a migração in-place, com backup
automático em ``data/output/_backup_pre_migracao_quem/``.

Idempotente: se o XLSX já está em formato genérico, só registra
no log e não escreve. Detecção via inspeção rápida da coluna
``quem`` da aba ``extrato``.

Uso:

    .venv/bin/python scripts/migrar_quem_generico.py [--executar]

Sem ``--executar``, faz dry-run (lista o que seria migrado).
"""

from __future__ import annotations

import argparse
import shutil
import sys
from pathlib import Path

import openpyxl

RAIZ = Path(__file__).resolve().parents[1]
DIR_OUTPUT = RAIZ / "data" / "output"
DIR_BACKUP = DIR_OUTPUT / "_backup_pre_migracao_quem"

MAPA_LEGACY: dict[str, str] = {
    "André": "pessoa_a",
    "Andre": "pessoa_a",
    "Vitória": "pessoa_b",
    "Vitoria": "pessoa_b",
    "Casal": "casal",
    "casal": "casal",
    "pessoa_a": "pessoa_a",
    "pessoa_b": "pessoa_b",
}


def normalizar_quem(valor: object) -> object:
    """Retorna valor normalizado se for legacy; senão, devolve intacto."""
    if not isinstance(valor, str):
        return valor
    if valor in MAPA_LEGACY:
        return MAPA_LEGACY[valor]
    return valor


def migrar_xlsx(arquivo: Path, dry_run: bool) -> dict[str, int]:
    """Aplica migração ao XLSX. Retorna estatísticas {abas, mudanças}."""
    estat = {"abas_visitadas": 0, "linhas_migradas": 0}
    wb = openpyxl.load_workbook(arquivo)
    mudou_global = False
    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        # Detecta coluna quem pelo cabeçalho.
        cabecalho = [
            (cell.value, cell.column) for cell in ws[1] if isinstance(cell.value, str)
        ]
        idx_quem = next(
            (col for valor, col in cabecalho if valor.strip().lower() == "quem"), None
        )
        if idx_quem is None:
            continue
        estat["abas_visitadas"] += 1
        for row in ws.iter_rows(min_row=2, min_col=idx_quem, max_col=idx_quem):
            for cell in row:
                novo = normalizar_quem(cell.value)
                if novo != cell.value:
                    cell.value = novo
                    estat["linhas_migradas"] += 1
                    mudou_global = True
    if mudou_global and not dry_run:
        wb.save(arquivo)
    return estat


def fazer_backup(arquivo: Path) -> Path:
    """Copia o XLSX para o diretório de backup. Retorna o caminho do backup."""
    DIR_BACKUP.mkdir(parents=True, exist_ok=True)
    destino = DIR_BACKUP / arquivo.name
    if not destino.exists():
        shutil.copy2(arquivo, destino)
    return destino


def main() -> int:
    parser = argparse.ArgumentParser(description="Migra coluna 'quem' dos XLSX.")
    parser.add_argument(
        "--executar", action="store_true", help="Aplica a migração (default: dry-run)."
    )
    args = parser.parse_args()
    dry = not args.executar

    if not DIR_OUTPUT.exists():
        print(f"data/output/ ausente em {DIR_OUTPUT}")
        return 1

    xlsx_files = sorted(DIR_OUTPUT.glob("*.xlsx"))
    if not xlsx_files:
        print("Nenhum XLSX em data/output/")
        return 0

    modo = "DRY-RUN" if dry else "EXECUTAR"
    print(f"[migrar_quem_generico] modo={modo} arquivos={len(xlsx_files)}")
    total_linhas = 0
    for arquivo in xlsx_files:
        if not dry:
            backup = fazer_backup(arquivo)
            print(f"  backup: {arquivo.name} -> {backup}")
        estat = migrar_xlsx(arquivo, dry_run=dry)
        total_linhas += estat["linhas_migradas"]
        print(
            f"  {arquivo.name}: abas={estat['abas_visitadas']} "
            f"linhas_migradas={estat['linhas_migradas']}"
        )

    if dry:
        print(
            f"[dry-run] {total_linhas} linha(s) seriam migradas. "
            "Rode com --executar para aplicar."
        )
    else:
        print(f"[ok] migração aplicada -- {total_linhas} linha(s) atualizada(s).")
    return 0


if __name__ == "__main__":
    sys.exit(main())


# "Quem teme migrar dado, vive escravo do legado." -- adágio do projeto
