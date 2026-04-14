"""Fase 4 do gauntlet: testa geração do XLSX com 8 abas."""

import tempfile
import time
from datetime import date
from pathlib import Path

import openpyxl

from scripts.gauntlet.config import (
    ABAS_XLSX_ESPERADAS,
    COLUNAS_EXTRATO,
    ResultadoFase,
    ResultadoTeste,
)
from src.utils.logger import configurar_logger

logger = configurar_logger("gauntlet.xlsx")


def _gerar_transacoes_sinteticas() -> list[dict]:
    """Gera transações normalizadas para teste do XLSX writer."""
    return [
        {
            "data": date(2026, 4, i + 1),
            "valor": float(100 * (i + 1)),
            "forma_pagamento": "Pix",
            "local": f"LOJA {i + 1}",
            "quem": "André" if i % 2 == 0 else "Vitória",
            "categoria": "Outros",
            "classificacao": "Questionável",
            "banco_origem": "Nubank",
            "tipo": "Despesa",
            "mes_ref": "2026-04",
            "tag_irpf": None,
            "obs": "",
        }
        for i in range(10)
    ]


def _testar_abas_existem(xlsx_path: Path) -> ResultadoTeste:
    """Verifica que o XLSX gerado tem as 8 abas obrigatórias."""
    inicio = time.time()

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        abas = wb.sheetnames
        wb.close()

        faltando = [a for a in ABAS_XLSX_ESPERADAS if a not in abas]
        passou = len(faltando) == 0

        detalhe = (
            f"{len(abas)} abas encontradas: {', '.join(abas)}"
            if passou
            else f"Faltando: {', '.join(faltando)}"
        )
    except Exception as e:
        passou = False
        detalhe = ""
        return ResultadoTeste(
            nome="abas_existem", passou=False,
            tempo=time.time() - inicio, erro=str(e),
        )

    return ResultadoTeste(
        nome="abas_existem",
        passou=passou,
        tempo=time.time() - inicio,
        detalhe=detalhe,
    )


def _testar_colunas_extrato(xlsx_path: Path) -> ResultadoTeste:
    """Verifica que a aba extrato tem as colunas corretas."""
    inicio = time.time()

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb["extrato"]
        cabecalho = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        wb.close()

        faltando = [c for c in COLUNAS_EXTRATO if c not in cabecalho]
        passou = len(faltando) == 0

        detalhe = (
            f"Colunas OK: {len(cabecalho)}"
            if passou
            else f"Faltando: {', '.join(faltando)}"
        )
    except Exception as e:
        passou = False
        detalhe = ""
        return ResultadoTeste(
            nome="colunas_extrato", passou=False,
            tempo=time.time() - inicio, erro=str(e),
        )

    return ResultadoTeste(
        nome="colunas_extrato",
        passou=passou,
        tempo=time.time() - inicio,
        detalhe=detalhe,
    )


def _testar_contagem_linhas(xlsx_path: Path, esperado: int) -> ResultadoTeste:
    """Verifica contagem de linhas na aba extrato."""
    inicio = time.time()

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb["extrato"]
        linhas = ws.max_row - 1 if ws.max_row and ws.max_row > 1 else 0
        wb.close()

        passou = linhas == esperado
        detalhe = f"{linhas} linhas (esperado: {esperado})"
    except Exception as e:
        passou = False
        detalhe = ""
        return ResultadoTeste(
            nome="contagem_linhas", passou=False,
            tempo=time.time() - inicio, erro=str(e),
        )

    return ResultadoTeste(
        nome="contagem_linhas",
        passou=passou,
        tempo=time.time() - inicio,
        detalhe=detalhe,
    )


def executar() -> ResultadoFase:
    """Gera XLSX com dados sintéticos e valida estrutura."""
    from src.load.xlsx_writer import gerar_xlsx

    fase = ResultadoFase(nome="xlsx")
    inicio = time.time()

    transacoes = _gerar_transacoes_sinteticas()

    with tempfile.TemporaryDirectory(prefix="gauntlet_xlsx_") as tmpdir:
        xlsx_path = Path(tmpdir) / "gauntlet_teste.xlsx"

        try:
            gerar_xlsx(transacoes, xlsx_path)
        except Exception as e:
            fase.testes.append(ResultadoTeste(
                nome="gerar_xlsx",
                passou=False,
                tempo=time.time() - inicio,
                erro=f"Falha ao gerar XLSX: {e}",
            ))
            fase.tempo_total = time.time() - inicio
            return fase

        fase.testes.append(ResultadoTeste(
            nome="gerar_xlsx",
            passou=True,
            tempo=0.0,
            detalhe="XLSX gerado sem erros",
        ))

        fase.testes.append(_testar_abas_existem(xlsx_path))
        fase.testes.append(_testar_colunas_extrato(xlsx_path))
        fase.testes.append(_testar_contagem_linhas(xlsx_path, len(transacoes)))

    fase.tempo_total = time.time() - inicio
    logger.info(
        "Fase xlsx: %d/%d testes OK em %.2fs",
        fase.ok, fase.total, fase.tempo_total,
    )
    return fase


# "Sem dados, você é apenas mais uma pessoa com uma opinião." -- W. Edwards Deming
