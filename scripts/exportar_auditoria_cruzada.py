"""Exporta XLSX de auditoria cruzada Opus × ETL × Graduação.

Sprint META-AUDITORIA-CRUZADA-XLSX (2026-05-16). Consolida em um único
XLSX o estado de cada arquivo já processado no projeto:

- O que o supervisor Opus principal leu via Read multimodal e gravou
  como prova artesanal canônica (``data/output/dossies/<tipo>/
  provas_artesanais/<sha256>.json``).
- O que o ETL extraiu e gravou como node ``documento`` no grafo
  (``data/output/grafo.sqlite``).
- O veredito automático do comparador (``data/output/dossies/<tipo>/
  comparacoes/<sha256>_<ts>.json``).
- O status global de graduação do tipo (``data/output/graduacao_tipos.json``).

Abas geradas:

1. **auditoria_cruzada** — uma linha por sha256 processado, com cruzamento
   Opus × ETL × Comparador.
2. **tipos_resumo** — uma linha por tipo canônico (22 do YAML), com
   métricas agregadas e gap analysis.
3. **divergencias_detalhe** — uma linha por divergência de campo (auditoria
   fina dos casos DIVERGENTE).
4. **outros_com_sugestao** — top transações "Outros" com sugestão TF-IDF
   do sugestor (apenas se ``sugestoes_categoria.json`` existe).
5. **stats_globais** — KPIs consolidados (uma linha).

Output: ``data/output/auditoria_cruzada_<YYYY-MM-DD>.xlsx``.

Uso:

    python -m scripts.exportar_auditoria_cruzada           # arquivo do dia
    python -m scripts.exportar_auditoria_cruzada --saida X # path custom
"""

from __future__ import annotations

import argparse
import json
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

_RAIZ = Path(__file__).resolve().parents[1]
DIR_DOSSIES = _RAIZ / "data" / "output" / "dossies"
PATH_GRAFO = _RAIZ / "data" / "output" / "grafo.sqlite"
PATH_GRADUACAO = _RAIZ / "data" / "output" / "graduacao_tipos.json"
PATH_TIPOS_YAML = _RAIZ / "mappings" / "tipos_documento.yaml"
PATH_SUGESTOES = _RAIZ / "data" / "output" / "sugestoes_categoria.json"
PATH_METRICAS = _RAIZ / "data" / "output" / "metricas_prontidao.json"


def _ler_json(path: Path) -> dict | list | None:
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None


def _resumir_campos_canonicos(campos: dict) -> str:
    """Resume campos canônicos da prova em string única (auditável)."""
    if not campos:
        return ""
    chaves_prioritarias = [
        "tipo_documento",
        "competencia",
        "data_emissao",
        "data_pagamento",
        "total",
        "valor",
        "liquido",
    ]
    partes = []
    for k in chaves_prioritarias:
        v = campos.get(k)
        if v is not None:
            partes.append(f"{k}={v}")
    # Estabelecimento / empresa / emitente:
    for sub in ("estabelecimento", "empresa", "emitente"):
        v = campos.get(sub)
        if isinstance(v, dict):
            nome = v.get("razao_social") or v.get("nome") or ""
            cnpj = v.get("cnpj") or ""
            if nome:
                partes.append(f"{sub}={nome[:40]}")
            if cnpj:
                partes.append(f"cnpj={cnpj}")
    return " | ".join(partes)


def _carregar_provas_e_comparacoes() -> dict[str, dict]:
    """Devolve dict {sha256: {tipo, prova, ultima_comparacao, ...}}.

    Apenas tipos com dossiê físico em ``data/output/dossies/``.
    """
    out: dict[str, dict] = {}
    if not DIR_DOSSIES.exists():
        return out
    for dir_tipo in sorted(DIR_DOSSIES.iterdir()):
        if not dir_tipo.is_dir():
            continue
        tipo = dir_tipo.name
        dir_provas = dir_tipo / "provas_artesanais"
        if not dir_provas.exists():
            continue
        for p in sorted(dir_provas.glob("*.json")):
            prova = _ler_json(p)
            if not isinstance(prova, dict):
                continue
            sha = prova.get("sha256") or p.stem
            out[sha] = {
                "tipo": tipo,
                "prova": prova,
                "lido_em": prova.get("lido_em", ""),
                "lido_por": prova.get("lido_por", ""),
                "campos": prova.get("campos_canonicos", {}),
                "notas": prova.get("_notas_supervisor", ""),
                "ultima_comparacao": None,
                "ultimo_veredito": "SEM_COMPARACAO",
            }
        # Última comparação por sha256 (timestamp mais recente):
        dir_comp = dir_tipo / "comparacoes"
        if dir_comp.exists():
            for c in sorted(dir_comp.glob("*.json")):
                comp = _ler_json(c)
                if not isinstance(comp, dict):
                    continue
                sha = comp.get("sha256")
                if sha and sha in out:
                    prev = out[sha].get("ultima_comparacao")
                    if prev is None or comp.get("comparado_em", "") > prev.get(
                        "comparado_em", ""
                    ):
                        out[sha]["ultima_comparacao"] = comp
                        out[sha]["ultimo_veredito"] = comp.get("veredito", "SEM_VEREDITO")
    return out


_REGEX_SHA8 = __import__("re").compile(
    r"_([0-9a-f]{8,16})\.(pdf|jpeg|jpg|png|xml)", __import__("re").IGNORECASE
)


def _extrair_sha_do_arquivo_origem(arquivo: str) -> str:
    """Heurística: extrai hash hex do nome do arquivo.

    Padrão canônico do projeto: ``<TIPO>_<data>_<sha8>.<ext>`` OU
    ``<TIPO>_<sha8>.<ext>``. Devolve o hash (até 16 chars) ou "" se
    não encontrar.
    """
    if not arquivo:
        return ""
    match = _REGEX_SHA8.search(arquivo)
    return match.group(1).lower() if match else ""


def _carregar_nodes_documento() -> dict[str, dict]:
    """Devolve dict indexado por sha8 OU sha completo do node.

    O grafo NÃO grava ``sha256`` completo em metadata — só ``arquivo_origem``
    com hash truncado embutido no nome (padrão ``<TIPO>_<sha8>.ext``).
    Para cruzar com provas (que têm sha256 completo), indexamos por sha8
    (primeiros 8-12 chars) extraído do arquivo_origem.
    """
    out: dict[str, dict] = {}
    if not PATH_GRAFO.exists():
        return out
    try:
        con = sqlite3.connect(str(PATH_GRAFO))
        try:
            cur = con.execute(
                "SELECT id, nome_canonico, metadata FROM node WHERE tipo='documento'"
            )
            for node_id, nome, meta_str in cur.fetchall():
                try:
                    meta = json.loads(meta_str) if meta_str else {}
                except json.JSONDecodeError:
                    meta = {}
                sha_completo = meta.get("sha256") or ""
                arquivo = meta.get("arquivo_origem", "")
                sha8 = _extrair_sha_do_arquivo_origem(arquivo)
                # Indexa por sha8 quando arquivo segue o padrão; senão por path:
                chave = sha8 if sha8 else (sha_completo if sha_completo else f"_path:{arquivo}")
                out[chave] = {
                    "node_id": node_id,
                    "nome_canonico": nome,
                    "metadata": meta,
                    "arquivo_origem": arquivo,
                    "sha256": sha_completo,
                    "sha8": sha8,
                }
        finally:
            con.close()
    except sqlite3.Error:
        return out
    return out


def _tipos_canonicos_do_yaml() -> list[str]:
    """Lê IDs dos 22 tipos canônicos. Fallback hardcoded se PyYAML ausente."""
    try:
        import yaml  # type: ignore[import-untyped]
    except ImportError:
        return []
    if not PATH_TIPOS_YAML.exists():
        return []
    try:
        doc = yaml.safe_load(PATH_TIPOS_YAML.read_text(encoding="utf-8")) or {}
    except yaml.YAMLError:
        return []
    return [t.get("id") for t in (doc.get("tipos") or []) if t.get("id")]


def _resumir_divergencias(comp: dict | None) -> str:
    if not comp or not isinstance(comp, dict):
        return ""
    divs = comp.get("divergencias", [])
    if not divs:
        return ""
    partes = []
    for d in divs[:5]:
        campo = d.get("campo", "?")
        partes.append(f"{campo}")
    if len(divs) > 5:
        partes.append(f"... (+{len(divs)-5})")
    return " | ".join(partes)


def _buscar_node_por_sha(
    sha_completo: str, nodes: dict[str, dict]
) -> dict | None:
    """Procura node compatível: testa sha completo, sha8 (primeiros 8 chars),
    sha12, depois fallback por path.
    """
    if not sha_completo:
        return None
    candidatos = [sha_completo, sha_completo[:12], sha_completo[:8]]
    for c in candidatos:
        if c in nodes:
            return nodes[c]
    return None


def montar_aba_auditoria_cruzada(
    provas: dict[str, dict], nodes: dict[str, dict]
) -> list[dict]:
    """Uma linha por arquivo (full outer join entre provas e nodes)."""
    # Coleta chaves canônicas (sha completo se possível):
    chaves_sha = set(provas.keys())
    sha8_para_node: dict[str, dict] = {n["sha8"]: n for n in nodes.values() if n.get("sha8")}

    # Nodes sem prova correspondente: chave = sha8 (não temos sha completo do ETL).
    # Para não duplicar, filtra nodes cujo sha8 bate algum prefixo de prova existente.
    sha8s_com_prova = set()
    for sha_prova in chaves_sha:
        sha8_prefix = sha_prova[:8] if sha_prova else ""
        if sha8_prefix in sha8_para_node:
            sha8s_com_prova.add(sha8_prefix)

    # Adiciona nodes órfãos (sem prova) usando sha8 como chave:
    for sha8, node in sha8_para_node.items():
        if sha8 not in sha8s_com_prova:
            chaves_sha.add(f"_orfao:{sha8}")

    linhas = []
    for sha in sorted(chaves_sha):
        if sha.startswith("_orfao:"):
            sha8 = sha[len("_orfao:"):]
            prova_info = None
            node_info = sha8_para_node.get(sha8)
            sha_display = sha8 + " (orfao)"
        else:
            prova_info = provas.get(sha)
            node_info = _buscar_node_por_sha(sha, nodes)
            sha_display = sha[:16] + "..." if len(sha) > 16 else sha

        tipo_opus = prova_info["tipo"] if prova_info else ""
        tipo_etl = ""
        if node_info:
            tipo_etl = node_info["metadata"].get("tipo_documento", "")

        linhas.append(
            {
                "sha256": sha_display,
                "sha256_completo": sha if not sha.startswith("_orfao:") else "",
                "tipo_opus": tipo_opus,
                "tipo_etl": tipo_etl,
                "tipo_match": tipo_opus == tipo_etl if tipo_opus and tipo_etl else "",
                "veredito_comparacao": (
                    prova_info["ultimo_veredito"] if prova_info else "SEM_PROVA"
                ),
                "opus_lido_em": prova_info["lido_em"] if prova_info else "",
                "opus_campos_resumo": (
                    _resumir_campos_canonicos(prova_info["campos"]) if prova_info else ""
                ),
                "etl_node_id": node_info["node_id"] if node_info else "",
                "etl_arquivo_origem": node_info["arquivo_origem"] if node_info else "",
                "etl_nome_canonico": node_info["nome_canonico"] if node_info else "",
                "n_divergencias": (
                    len((prova_info.get("ultima_comparacao") or {}).get("divergencias", []))
                    if prova_info and prova_info.get("ultima_comparacao")
                    else 0
                ),
                "divergencias_resumo": _resumir_divergencias(
                    prova_info.get("ultima_comparacao") if prova_info else None
                ),
                "opus_notas_supervisor": (
                    prova_info["notas"][:200] if prova_info else ""
                ),
            }
        )
    return linhas


def montar_aba_tipos_resumo(
    tipos_canonicos: list[str],
    graduacao: dict | None,
    provas: dict[str, dict],
    nodes: dict[str, dict],
) -> list[dict]:
    """Uma linha por tipo canônico do YAML, mais 'sem_dossie' para órfãos."""
    grad_tipos = (graduacao or {}).get("tipos", {}) if isinstance(graduacao, dict) else {}

    # Contadores derivados:
    provas_por_tipo: dict[str, int] = {}
    nodes_por_tipo: dict[str, int] = {}
    for info in provas.values():
        provas_por_tipo[info["tipo"]] = provas_por_tipo.get(info["tipo"], 0) + 1
    for node in nodes.values():
        tipo = node["metadata"].get("tipo_documento", "")
        if tipo:
            nodes_por_tipo[tipo] = nodes_por_tipo.get(tipo, 0) + 1

    todos_tipos = set(tipos_canonicos) | set(grad_tipos.keys()) | set(provas_por_tipo.keys())
    linhas = []
    for tipo in sorted(todos_tipos):
        grad_info = grad_tipos.get(tipo, {})
        amostras_ok = grad_info.get("amostras_ok", 0)
        if isinstance(amostras_ok, list):
            amostras_ok = len(amostras_ok)
        amostras_div = grad_info.get("amostras_divergentes", 0)
        if isinstance(amostras_div, list):
            amostras_div = len(amostras_div)
        linhas.append(
            {
                "tipo": tipo,
                "esta_no_yaml_canonico": tipo in tipos_canonicos,
                "status_graduacao": grad_info.get("status", "SEM_DOSSIE"),
                "amostras_ok": amostras_ok,
                "amostras_divergentes": amostras_div,
                "provas_artesanais": provas_por_tipo.get(tipo, 0),
                "nodes_no_grafo": nodes_por_tipo.get(tipo, 0),
                "atualizado_em": grad_info.get("atualizado_em", ""),
            }
        )
    return linhas


def montar_aba_divergencias_detalhe(provas: dict[str, dict]) -> list[dict]:
    """Uma linha por divergência de campo. Permite auditoria fina."""
    linhas = []
    for sha, info in provas.items():
        comp = info.get("ultima_comparacao")
        if not comp:
            continue
        for d in comp.get("divergencias", []):
            linhas.append(
                {
                    "sha256": sha[:16] + "..." if len(sha) > 16 else sha,
                    "tipo": info["tipo"],
                    "campo": d.get("campo", ""),
                    "tipo_divergencia": d.get("tipo", ""),
                    "esperado_opus": str(d.get("esperado", ""))[:100],
                    "obtido_etl": str(d.get("obtido", ""))[:100],
                    "comparado_em": comp.get("comparado_em", ""),
                }
            )
    return linhas


def montar_aba_outros_com_sugestao(top_n: int = 100) -> list[dict]:
    """Top N transações "Outros" com sugestão TF-IDF (do sugestor)."""
    sug = _ler_json(PATH_SUGESTOES)
    if not isinstance(sug, dict):
        return []
    sugestoes = sug.get("sugestoes", {})
    linhas = []
    # Ordena por confiança decrescente:
    items = sorted(
        sugestoes.items(),
        key=lambda kv: -float(kv[1].get("confianca_top1", 0)),
    )
    for tx_id, item in items[:top_n]:
        linhas.append(
            {
                "id_transacao": tx_id,
                "descricao": (item.get("descricao", "") or "")[:80],
                "categoria_atual": "Outros",
                "sugestao_top1": item.get("top1", ""),
                "confianca_top1": item.get("confianca_top1", 0),
                "n_sugestoes": len(item.get("sugestoes", [])),
            }
        )
    return linhas


def montar_aba_amostras_faltantes(
    tipos_canonicos: list[str],
    graduacao: dict | None,
    provas: dict[str, dict],
) -> list[dict]:
    """Lista tipos que precisam de coleta humana para chegar à meta ≥15 GRADUADOS.

    Para cada tipo PENDENTE ou SEM_DOSSIE no YAML:
    - tipo_id, descricao_humanizada
    - amostras_faltantes (até 2 para graduar — meta inicial)
    - prioridade_coleta (alta/média/baixa baseado em probabilidade
      do dono ter o documento)
    - exemplo_de_pasta (onde dono pode procurar)
    """
    grad_tipos = (graduacao or {}).get("tipos", {}) if isinstance(graduacao, dict) else {}

    # Probabilidade do dono ter o documento (curadoria manual baseada em
    # FASE-A-AGUARDA-AMOSTRAS-2026-05-14):
    prioridade: dict[str, tuple[str, str, str]] = {
        # tipo_id: (prioridade, descricao_humana, exemplo_pasta)
        "das_mei": (
            "alta",
            "DAS do MEI (mensal). Andre tem MEI ativo.",
            "Receita Federal / app MEI",
        ),
        "irpf_parcela": (
            "alta",
            "Parcela do IRPF (DARF emitido pela DIRPF).",
            "Receita Federal / app DIRPF",
        ),
        "comprovante_cpf": (
            "média",
            "Comprovante CPF (1 vez na vida).",
            "Receita Federal / app",
        ),
        "certidao_receita_cnpj": (
            "média",
            "Certidão negativa de débitos do CNPJ MEI.",
            "Receita Federal / app",
        ),
        "conta_luz": (
            "alta",
            "Conta de energia Neoenergia DF (mensal).",
            "app Neoenergia / email",
        ),
        "conta_agua": (
            "alta",
            "Conta de água CAESB (mensal).",
            "app CAESB / email",
        ),
        "recibo_nao_fiscal": (
            "média",
            "Recibo de serviço sem nota fiscal (ex: serviços eventuais).",
            "celular / impressos",
        ),
        "receita_medica": (
            "média",
            "Receita médica (dedutível IRPF se for de saúde).",
            "celular / consultórios",
        ),
        "garantia_fabricante": (
            "baixa",
            "Termo de garantia do fabricante (≠ cupom_garantia_estendida).",
            "manuais de produtos comprados",
        ),
        "contrato": (
            "baixa",
            "Contrato genérico (locação, serviços).",
            "arquivo doméstico",
        ),
        "danfe_nfe55": (
            "baixa",
            "NFE modelo 55 (DANFE) - notas eletrônicas full.",
            "compras online / fornecedores",
        ),
        "xml_nfe": (
            "baixa",
            "XML da NFE (recebimento por email após compra).",
            "email após compra B2B",
        ),
        "extrato_c6_pdf": (
            "média",
            "Extrato C6 em PDF (alternativo ao XLSX/OFX). Já há extrato_bancario GRADUADO.",
            "app C6 export",
        ),
    }
    linhas = []
    for tipo in sorted(tipos_canonicos):
        grad_info = grad_tipos.get(tipo, {})
        status = grad_info.get("status", "SEM_DOSSIE")
        if status == "GRADUADO":
            continue
        amostras_ok = grad_info.get("amostras_ok", 0)
        if isinstance(amostras_ok, list):
            amostras_ok = len(amostras_ok)
        faltam = max(2 - amostras_ok, 0)
        pri, desc, exemplo = prioridade.get(tipo, ("baixa", "", ""))
        linhas.append(
            {
                "tipo": tipo,
                "status_atual": status,
                "amostras_atuais": amostras_ok,
                "amostras_faltantes": faltam,
                "prioridade_coleta": pri,
                "descricao": desc,
                "onde_buscar": exemplo,
            }
        )
    return linhas


def montar_aba_stats_globais() -> list[dict]:
    """KPIs consolidados em uma linha."""
    m = _ler_json(PATH_METRICAS)
    if not isinstance(m, dict):
        return []
    return [
        {"metrica": "gerado_em", "valor": m.get("gerado_em", "")},
        {"metrica": "tipos_graduados", "valor": m.get("tipos_graduados", 0)},
        {"metrica": "tipos_total_canonico", "valor": m.get("tipos_total_canonico", 22)},
        {
            "metrica": "linking_documento_de_pct (sobre transacoes)",
            "valor": m.get("linking_documento_de_pct", 0),
        },
        {
            "metrica": "linking_documento_de_linked",
            "valor": m.get("linking_documento_de_linked", 0),
        },
        {
            "metrica": "linking_documento_de_total_transacoes",
            "valor": m.get("linking_documento_de_total_transacoes", 0),
        },
        {
            "metrica": "categorizacao_outros_pct",
            "valor": m.get("categorizacao_outros_pct", 0),
        },
        {
            "metrica": "categorizacao_outros_count",
            "valor": m.get("categorizacao_outros_count", 0),
        },
        {"metrica": "pytest_count", "valor": m.get("pytest_count", 0)},
        {
            "metrica": "backup_grafo_automatico",
            "valor": m.get("backup_grafo_automatico", False),
        },
        {
            "metrica": "transacionalidade_pipeline",
            "valor": m.get("transacionalidade_pipeline", False),
        },
        {
            "metrica": "lockfile_concorrencia",
            "valor": m.get("lockfile_concorrencia", False),
        },
    ]


def gerar_xlsx(saida: Path) -> dict:
    """Pipeline completo: coleta dados, monta abas, grava XLSX."""
    try:
        import openpyxl  # type: ignore[import-untyped]
    except ImportError as exc:
        raise RuntimeError("openpyxl ausente; install via pip") from exc

    provas = _carregar_provas_e_comparacoes()
    nodes = _carregar_nodes_documento()
    tipos = _tipos_canonicos_do_yaml()
    graduacao = _ler_json(PATH_GRADUACAO)

    abas = {
        "auditoria_cruzada": montar_aba_auditoria_cruzada(provas, nodes),
        "tipos_resumo": montar_aba_tipos_resumo(tipos, graduacao, provas, nodes),
        "amostras_faltantes": montar_aba_amostras_faltantes(tipos, graduacao, provas),
        "divergencias_detalhe": montar_aba_divergencias_detalhe(provas),
        "outros_com_sugestao": montar_aba_outros_com_sugestao(top_n=100),
        "stats_globais": montar_aba_stats_globais(),
    }

    wb = openpyxl.Workbook()
    # Remove default sheet:
    wb.remove(wb.active)
    for nome_aba, linhas in abas.items():
        ws = wb.create_sheet(title=nome_aba)
        if not linhas:
            ws.append(["(sem dados)"])
            continue
        # Header = chaves do primeiro dict:
        headers = list(linhas[0].keys())
        ws.append(headers)
        for linha in linhas:
            ws.append([linha.get(h, "") for h in headers])
        # Bold header:
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

    saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(saida))
    return {
        "saida": str(saida),
        "abas": {nome: len(linhas) for nome, linhas in abas.items()},
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Auditoria cruzada Opus × ETL")
    parser.add_argument(
        "--saida",
        type=Path,
        default=None,
        help="Path do XLSX (default: data/output/auditoria_cruzada_<YYYY-MM-DD>.xlsx)",
    )
    args = parser.parse_args(argv)

    if args.saida is None:
        ts = datetime.now().strftime("%Y-%m-%d")
        args.saida = _RAIZ / "data" / "output" / f"auditoria_cruzada_{ts}.xlsx"

    info = gerar_xlsx(args.saida)
    sys.stdout.write(f"Auditoria cruzada em {info['saida']}\n")
    for nome, n in info["abas"].items():
        sys.stdout.write(f"  {nome}: {n} linhas\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())


# "Auditoria honesta cruza o que duas testemunhas dizem." -- principio
